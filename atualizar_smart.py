# -*- coding: utf-8 -*-
"""
Atualizador Dashboard Firjan - Campanha Smart Factory
Le os arquivos Mailing e Retorno e atualiza automaticamente o index.html

Mailing : empresas (CNPJs distintos, coluna E)
Retorno : tentativas, status, evolucao diaria
          aba: chamadas_22-04-2026_165407
          status: coluna L (STATUS_NEGOCIO); se vazia, usa coluna K (STATUS)
"""

import openpyxl
import re
import glob
import os
import io
import urllib.request
import http.cookiejar
from collections import Counter, defaultdict
from datetime import datetime, timedelta

# ═══════════════════════════════════════════════════════════
# CONFIGURAÇÃO
# ═══════════════════════════════════════════════════════════

PASTA_SMART   = r'Arquivos\nao_atualizaveis\Ativo_Smart_Factory'
ABA_RETORNO   = 'chamadas_22-04-2026_165407'
INDEX_HTML    = r'index.html'
DEPARA_PATH   = r'Arquivos\bases_apoio\tab_de-para.xlsx'

# --- Agendamentos Smart Factory (Discagem) ---
PASTA_AGEND   = r'Arquivos\nao_atualizaveis\Agendamento_Smart_Factory'
PREFIXO_AGEND = 'Discagem_Smart Factory Agendamentos'

# --- Lista de Agendamentos Smart Factory (agenda propriamente dita, SharePoint) ---
# Colunas: E(4) RAZAO SOCIAL, I(8) Solucao de Interesse, L(11) DATA E HORA AGENDAMENTO
AGEND_LISTA_URL = ('https://ddmadvbr-my.sharepoint.com/:x:/g/personal/'
                   'denisesantos_firjan_grupoddm_com_br/'
                   'IQDEsPo9BeR-QIfKcR0znz_eAW6NK3xj6Xb6VJWtP3SDJE0?download=1')
AL_RAZAO = 4   # E
AL_SOL   = 8   # I
AL_DATA  = 11  # L

# Número da empresa — quando aparece em ORIGEM, usar DESTINO
NOSSO_NUMERO = '2120384382'  # (21) 2038-4382 normalizado

# Decisor e Interessado determinados pelo LABEL do de-para:
# Decisor    = label não está em LABELS_NAO_DECISOR
# Interessado = label == 'Interessado'
LABELS_NAO_DECISOR = {'Telefonia', 'Tentativa', 'Engano', 'Alo'}
LABEL_INTERESSADO  = 'Interessado'

# Agendamentos: status (raw, nao esta no de-para) que indica agendamento concluido
LABEL_AGENDADO = 'Agendamento Realizado'

# Agendamentos: tabulacoes brutas (antes do de-para) que contam como "Contatos com Decisor"
TAB_AGEND_DECISOR = {'AGENDAMENTO REALIZADO', 'NAO INTERESSADO', 'INTERESSADO', 'CLIENTE DESLIGOU'}

STATUS_MAP = {}

MES_PT = {1:'Jan',2:'Fev',3:'Mar',4:'Abr',5:'Mai',6:'Jun',
          7:'Jul',8:'Ago',9:'Set',10:'Out',11:'Nov',12:'Dez'}

# ═══════════════════════════════════════════════════════════
# FUNÇÕES AUXILIARES
# ═══════════════════════════════════════════════════════════

def fmt_num(n):
    return f"{n:,}".replace(",", ".")

def fmt_pct(n):
    return f"{n:.2f}%".replace(".", ",")

def fmt_dec(n):
    return f"{n:.2f}".replace(".", ",")

def norm_tel(t):
    return re.sub(r'\D', '', str(t)) if t else ''

def parse_data_agendamento(val):
    """Converte a celula DATA E HORA AGENDAMENTO em datetime.
    Aceita datetime/serial do Excel OU texto livre digitado manualmente,
    nos formatos 'DD/MM/AAAA HH:MM' ou 'DD/MM HH:MM [horas]' (sem ano -> 2026)."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, (int, float)) and val > 0:
        try:
            return datetime(1899, 12, 30) + timedelta(days=float(val))
        except Exception:
            return None
    if isinstance(val, str):
        s = val.strip()
        m = re.match(r'(\d{1,2})/(\d{1,2})/(\d{4})\s+(\d{1,2}):(\d{2})', s)
        if m:
            dd, mo, yy, hh, mi = m.groups()
            try:
                return datetime(int(yy), int(mo), int(dd), int(hh), int(mi))
            except Exception:
                return None
        m = re.match(r'(\d{1,2})/(\d{1,2})\s+(\d{1,2}):(\d{2})', s)
        if m:
            dd, mo, hh, mi = m.groups()
            try:
                return datetime(2026, int(mo), int(dd), int(hh), int(mi))
            except Exception:
                return None
    return None

def baixar_sharepoint(url):
    """Baixa o .xlsx do SharePoint (link anonimo) via cookie jar."""
    cj = http.cookiejar.CookieJar()
    op = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(cj))
    op.addheaders = [('User-Agent', 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)')]
    resp = op.open(url, timeout=60)
    return resp.read()

def norm_txt(s):
    """Maiuscula, sem acento - para comparar tabulacoes brutas."""
    import unicodedata
    b = unicodedata.normalize('NFKD', str(s).strip().upper())
    return ''.join(c for c in b if not unicodedata.combining(c))

def to_datetime(val):
    if val is None: return None
    if isinstance(val, datetime): return val
    if isinstance(val, (int, float)) and val > 0:
        try: return datetime(1899, 12, 30) + timedelta(days=float(val))
        except: return None
    return None

def ler_depara(caminho):
    mapa = {}
    try:
        wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
        ws = wb['status'] if 'status' in wb.sheetnames else wb.active  # aba "status"
        first = True
        for row in ws.iter_rows(values_only=True):
            if first: first = False; continue
            if row[0] and row[1]:
                mapa[str(row[0]).strip()] = str(row[1]).strip()
        wb.close()
        print(f'  De-para: {len(mapa)} entradas')
    except FileNotFoundError:
        print(f'  [AVISO] De-para nao encontrado: {caminho}')
    return mapa

def normalizar_status(raw):
    if raw is None: return None
    s = str(raw).strip()
    for k, v in STATUS_MAP.items():
        if s.upper() == k.upper(): return v
    return s

def encontrar_arquivo(pasta, prefixo):
    padrao = os.path.join(pasta, f'{prefixo}*.xlsx')
    arquivos = [a for a in glob.glob(padrao) if not os.path.basename(a).startswith('~$')]
    if not arquivos:
        raise FileNotFoundError(f'\n[ERRO] Arquivo nao encontrado: {prefixo}*.xlsx em {pasta}')
    return sorted(arquivos)[-1]

# ═══════════════════════════════════════════════════════════
# LEITURA
# ═══════════════════════════════════════════════════════════

def calcular_mailing_smart(caminho):
    """Lê o Mailing Smart Factory: conta CNPJs distintos da coluna E (idx 4)."""
    print(f'  Lendo Mailing: {os.path.basename(caminho)}')
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb.active
    cnpjs = set()
    tels  = set()
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0: continue  # pula header
        if row[4]: cnpjs.add(row[4])   # coluna E = CNPJ
        if row[7]: tels.add(norm_tel(row[7]))  # coluna H = TELEFONE
    wb.close()
    return {'_empresas': len(cnpjs), '_tels_mailing': tels}


def calcular_retorno_smart(caminho, aba):
    """Lê o Retorno Smart Factory.
    Cada linha com DATA = 1 tentativa.
    Status: col L (STATUS_NEGOCIO); se vazia, usa col K (STATUS).
    """
    print(f'  Lendo Retorno: {os.path.basename(caminho)} | aba: {aba}')
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb[aba]
    headers = [str(h) if h else '' for h in next(ws.iter_rows(values_only=True))]
    wb_rows = list(ws.iter_rows(values_only=True, min_row=2))
    wb.close()

    # Índices das colunas-chave
    data_idx = 0   # col A = DATA
    orig_idx = 7   # col H = ORIGEM
    dest_idx = 8   # col I = DESTINO
    tipo_idx = 9   # col J = TIPO
    st_idx   = 10  # col K = STATUS
    sn_idx   = 11  # col L = STATUS_NEGOCIO

    total_tent    = 0
    status_counter = Counter()
    raw_por_label  = defaultdict(set)
    evolucao      = {}
    tel_decisor   = set()
    tel_interesse = set()

    for row in wb_rows:
        dt = to_datetime(row[data_idx])
        if not dt:
            continue  # só conta linha com data

        total_tent += 1

        # Telefone: se ORIGEM for o nosso número → usar DESTINO; senão → ORIGEM
        orig_norm = norm_tel(row[orig_idx])
        if orig_norm == NOSSO_NUMERO:
            tel = norm_tel(row[dest_idx])
        else:
            tel = orig_norm

        # Status: usa L se preenchido, senão K
        sn_raw = row[sn_idx]
        st_raw = row[st_idx]
        raw    = str(sn_raw).strip() if sn_raw else (str(st_raw).strip() if st_raw else '')
        label  = normalizar_status(raw) if raw else None

        if label:
            status_counter[label] += 1
            if raw: raw_por_label[label].add(raw)

        # Decisor = label preenchido e não está em LABELS_NAO_DECISOR
        if label and label not in LABELS_NAO_DECISOR:
            tel_decisor.add(tel)

        # Interessado = label == 'Interessado'
        if label == LABEL_INTERESSADO:
            tel_interesse.add(tel)

        # Evolução diária
        dk = f"{dt.day:02d}/{MES_PT[dt.month]}"
        if dk not in evolucao:
            evolucao[dk] = {'date': dt.date(), 'tent': 0, 'conv': 0}
        evolucao[dk]['tent'] += 1
        if label == LABEL_INTERESSADO:
            evolucao[dk]['conv'] += 1

    return {
        '_tentativas':    total_tent,
        '_statusCounter': status_counter,
        '_raw_por_label': dict(raw_por_label),
        '_evolucao':      evolucao,
        '_tel_decisor':   tel_decisor,
        '_tel_interesse': tel_interesse,
    }


def calcular_wpp_smart(caminho):
    """Lê o Retorno_Smart_Factory_Whatsapp.xlsx.
    Aba 'Enviados'  : cada linha (sem header) = 1 mensagem enviada.
    Aba 'Respostas' : cada linha (sem header) = 1 resposta; col G (idx 6) = tabulação.
    """
    print(f'  Lendo WhatsApp: {os.path.basename(caminho)}')
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)

    # --- Enviados ---
    ws_env = wb['Enviados']
    enviados = 0
    for row in ws_env.iter_rows(values_only=True, min_row=2):  # min_row=2 pula cabeçalho
        if any(c for c in row):
            enviados += 1

    # --- Respostas e tabulações ---
    ws_res = wb['Respostas']
    respostas = 0
    tab_counter = Counter()
    for i, row in enumerate(ws_res.iter_rows(values_only=True)):
        if i == 0: continue  # header
        if any(c for c in row):
            respostas += 1
            tab = row[6]  # coluna G
            if tab:
                tab_counter[str(tab).strip()] += 1

    wb.close()

    sem_resp = enviados - respostas
    taxa     = (respostas / enviados * 100) if enviados > 0 else 0

    # Tabulações relevantes para o resumo
    informado = tab_counter.get('Informado', 0)
    email     = tab_counter.get('Enviar E-mail', 0)

    print(f'    Enviados : {enviados}  |  Respostas: {respostas}  |  Sem Resp: {sem_resp}')
    print(f'    Taxa     : {taxa:.2f}%  |  Informado: {informado}  |  E-mail: {email}')

    return {
        '_wpp_enviados':  enviados,
        '_wpp_respostas': respostas,
        '_wpp_sem':       sem_resp,
        '_wpp_taxa':      taxa,
        '_wpp_informado': informado,
        '_wpp_email':     email,
    }


def calcular_agendamentos_smart(caminho):
    """Le o Discagem_Smart Factory Agendamentos.xlsx.
    Cada linha com DATA (col A) = 1 ligacao de agendamento.
    Status: col L (STATUS_NEGOCIO); se vazia, usa col K (STATUS).
    """
    print(f'  Lendo Agendamentos: {os.path.basename(caminho)}')
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    aba = next((s for s in wb.sheetnames if s.startswith('chamadas_')), wb.sheetnames[0])
    print(f'    aba: {aba}')
    ws = wb[aba]

    data_idx = 0   # col A = DATA
    orig_idx = 7   # col H = ORIGEM
    dest_idx = 8   # col I = DESTINO
    st_idx   = 10  # col K = STATUS
    sn_idx   = 11  # col L = STATUS_NEGOCIO

    total          = 0
    status_counter = Counter()
    raw_por_label  = defaultdict(set)
    evolucao       = {}
    tel_decisor    = set()

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        dt = to_datetime(row[data_idx])
        if not dt:
            continue
        total += 1

        sn_raw = row[sn_idx] if len(row) > sn_idx else None
        st_raw = row[st_idx] if len(row) > st_idx else None
        raw    = str(sn_raw).strip() if sn_raw else (str(st_raw).strip() if st_raw else '')
        label  = normalizar_status(raw) if raw else None
        rawn   = norm_txt(raw) if raw else ''

        if label:
            status_counter[label] += 1
            if raw: raw_por_label[label].add(raw)

        if rawn in TAB_AGEND_DECISOR:
            orig_norm = norm_tel(row[orig_idx])
            tel = norm_tel(row[dest_idx]) if orig_norm == NOSSO_NUMERO else orig_norm
            tel_decisor.add(tel)

        dk = f"{dt.day:02d}/{MES_PT[dt.month]}"
        if dk not in evolucao:
            evolucao[dk] = {'date': dt.date(), 'qtd': 0, 'agendadas': 0}
        evolucao[dk]['qtd'] += 1
        if label == LABEL_AGENDADO:
            evolucao[dk]['agendadas'] += 1

    wb.close()

    dias_ord = sorted(evolucao.items(), key=lambda x: x[1]['date'])
    st_items = status_counter.most_common()
    st_tooltips = [', '.join(sorted(raw_por_label.get(s[0], set()))) for s in st_items]

    print(f'  Total Agendamentos: {total}')
    print(f'  Contatos com Decisor: {len(tel_decisor)}')
    print(f'  Status: {dict(st_items[:5])} ...')

    return {
        'agendTotal':          total,
        'agendDecisor':        len(tel_decisor),
        'agendStatusLabels':   [s[0] for s in st_items],
        'agendStatusData':     [s[1] for s in st_items],
        'agendStatusTooltips': st_tooltips,
        'agendEvolucaoLabels': [d[0] for d in dias_ord],
        'agendTentDia':        [d[1]['qtd'] for d in dias_ord],
        'agendConvDia':        [d[1]['agendadas'] for d in dias_ord],
    }


def calcular_lista_agendamentos_sharepoint():
    """Baixa a Lista de Agendamentos Smart Factory do SharePoint.
    Colunas: E(4) RAZAO SOCIAL, I(8) Solucao de Interesse, L(11) DATA E HORA AGENDAMENTO.
    Empresas na Base = Razao Social unicas (todas as linhas).
    Agendamentos = linhas com a coluna L preenchida (data e hora marcada).
    """
    print('  Baixando Lista de Agendamentos Smart Factory (SharePoint)...')
    dados = baixar_sharepoint(AGEND_LISTA_URL)
    wb = openpyxl.load_workbook(io.BytesIO(dados), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    razoes = set()
    linhas_marcadas = []

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if not any(c is not None and str(c).strip() for c in row[:13]):
            continue
        if len(row) > AL_RAZAO and row[AL_RAZAO]:
            razoes.add(str(row[AL_RAZAO]).strip())

        dt = parse_data_agendamento(row[AL_DATA]) if len(row) > AL_DATA else None
        if dt:
            razao = str(row[AL_RAZAO]).strip() if len(row) > AL_RAZAO and row[AL_RAZAO] else ''
            sol   = str(row[AL_SOL]).strip()   if len(row) > AL_SOL   and row[AL_SOL]   else ''
            linhas_marcadas.append((dt, razao, sol))

    wb.close()
    linhas_marcadas.sort(key=lambda x: x[0])

    empresas_com_agend = {razao for _, razao, _ in linhas_marcadas if razao}

    print(f'  Empresas na base: {len(razoes)}')
    print(f'  Agendamentos (coluna L preenchida): {len(linhas_marcadas)} | Empresas com agendamento: {len(empresas_com_agend)}')

    return {
        'empresas':          len(razoes),
        'agendamentos':      len(linhas_marcadas),
        'agendamentosEmpresas': len(empresas_com_agend),
        'agendListaData':  [f"{dt.day:02d}/{MES_PT[dt.month]} {dt.hour:02d}:{dt.minute:02d}" for dt, _, _ in linhas_marcadas],
        'agendListaRazao': [razao for _, razao, _ in linhas_marcadas],
        'agendListaSol':   [sol for _, _, sol in linhas_marcadas],
    }


def combinar_smart(m, r, w=None):
    """Combina Mailing + Retorno Smart Factory (+ WhatsApp opcional)."""
    empresas  = m['_empresas']
    tent      = r['_tentativas']
    decisor   = len(r['_tel_decisor'])
    interesse = len(r['_tel_interesse'])
    taxa  = (interesse / decisor * 100) if decisor > 0 else 0
    media = tent / empresas if empresas > 0 else 0

    dias_ord = sorted(r['_evolucao'].items(), key=lambda x: x[1]['date'] or datetime(9999,12,31).date())
    st_items = r['_statusCounter'].most_common()
    rpl      = r.get('_raw_por_label', {})
    st_tooltips = [', '.join(sorted(rpl.get(s[0], set()))) for s in st_items]

    # WhatsApp
    show_wpp  = w is not None
    wpp_env   = fmt_num(w['_wpp_enviados'])  if w else '-'
    wpp_resp  = fmt_num(w['_wpp_respostas']) if w else '-'
    wpp_taxa  = fmt_pct(w['_wpp_taxa'])      if w else '-'
    wpp_sem   = fmt_num(w['_wpp_sem'])       if w else '-'
    wpp_info  = fmt_num(w['_wpp_informado']) if w else '-'
    wpp_email = fmt_num(w['_wpp_email'])     if w else '-'
    wpp_pie   = [w['_wpp_respostas'], w['_wpp_sem']] if w else [0, 1]

    return {
        '_empresas':      empresas,
        '_tentativas':    tent,
        '_interessados':  interesse,
        '_decisor':       decisor,
        '_statusCounter': r['_statusCounter'],
        '_raw_por_label': rpl,
        '_evolucao':      r['_evolucao'],
        'empresas':      fmt_num(empresas),
        'tentativas':    fmt_num(tent),
        'interessados':  fmt_num(interesse),
        'conversao':     fmt_pct(taxa),
        'decisor':       fmt_num(decisor),
        'media':         fmt_dec(media),
        'statusLabels':  [s[0] for s in st_items],
        'statusData':    [s[1] for s in st_items],
        'statusTooltips': st_tooltips,
        'evoLabels':     [d[0] for d in dias_ord],
        'tentDia':       [d[1]['tent'] for d in dias_ord],
        'convDia':       [d[1]['conv'] for d in dias_ord],
        'showWpp':       show_wpp,
        'wppEnv':        wpp_env,
        'wppResp':       wpp_resp,
        'wppTaxa':       wpp_taxa,
        'wppSem':        wpp_sem,
        'wppInfo':       wpp_info,
        'wppEmail':      wpp_email,
        'wppPie':        wpp_pie,
    }


def calcular_kpis_smart():
    m = calcular_mailing_smart(encontrar_arquivo(PASTA_SMART, 'Mailing_Smart_Factory'))
    r = calcular_retorno_smart(encontrar_arquivo(PASTA_SMART, 'Retorno_Ativo_Smart_Factory'), ABA_RETORNO)
    try:
        w = calcular_wpp_smart(encontrar_arquivo(PASTA_SMART, 'Retorno_Smart_Factory_Whatsapp'))
    except FileNotFoundError as e:
        print(f'  [AVISO] WhatsApp nao encontrado — secao sera ocultada. {e}')
        w = None
    return combinar_smart(m, r, w)


def calcular_kpis_smart_agend():
    """Campanha 'Smart Factory - Agendamentos' (aba propria)."""
    lista   = calcular_lista_agendamentos_sharepoint()
    agend   = calcular_agendamentos_smart(encontrar_arquivo(PASTA_AGEND, PREFIXO_AGEND))

    empresas = lista['empresas']
    agendamentos_total     = lista['agendamentos']
    agendamentos_empresas  = lista['agendamentosEmpresas']
    tent    = agend['agendTotal']
    decisor = agend['agendDecisor']
    media   = (tent / empresas) if empresas > 0 else 0
    taxa    = (agendamentos_empresas / decisor * 100) if decisor > 0 else 0

    return {
        'empresas':      fmt_num(empresas),
        'tentativas':    fmt_num(tent),
        'decisor':       fmt_num(decisor),
        'media':         fmt_dec(media),
        'agendamentos':       fmt_num(agendamentos_empresas),
        'agendamentosTotal':  fmt_num(agendamentos_total),
        'conversao':     fmt_pct(taxa),
        'statusLabels':  agend['agendStatusLabels'],
        'statusData':    agend['agendStatusData'],
        'statusTooltips': agend['agendStatusTooltips'],
        'evoLabels':     agend['agendEvolucaoLabels'],
        'tentDia':       agend['agendTentDia'],
        'convDia':       agend['agendConvDia'],
        'agendListaData':  lista['agendListaData'],
        'agendListaRazao': lista['agendListaRazao'],
        'agendListaSol':   lista['agendListaSol'],
    }

# ═══════════════════════════════════════════════════════════
# GERAÇÃO DO BLOCO JS
# ═══════════════════════════════════════════════════════════

def gerar_bloco_smart(k):
    def js_str(lst): return '[' + ','.join(f"'{str(v).replace(chr(39), chr(92)+chr(39))}'" for v in lst) + ']'
    def js_num(lst): return '[' + ','.join(str(v) for v in lst) + ']'

    show_wpp  = 'true' if k.get('showWpp') else 'false'
    wpp_pie   = js_num(k.get('wppPie', [0, 1]))

    periodo = f"{k['evoLabels'][0]} — {k['evoLabels'][-1]}" if k['evoLabels'] else ''
    return f"""  /* SMART_START */
  smart: {{
    label: '— Smart Factory', desc: 'Campanha Smart Factory — dados filtrados', periodo: '{periodo}',
    empresas: '{k['empresas']}', empresasLabel: '🏢 Empresas na Base',
    mediaLabel: '🔁 Média Tentativas/Empresa', mediaSub: 'por empresa',
    tentativas: '{k['tentativas']}', interessados: '{k['interessados']}', conversao: '{k['conversao']}',
    decisor: '{k['decisor']}', decisorSub: 'Apenas Smart Factory', media: '{k['media']}', trend: '',
    statusLabels: {js_str(k['statusLabels'])},
    statusData: {js_num(k['statusData'])}, statusColors:null,
    statusTooltips: {js_str(k.get('statusTooltips', ['']*len(k['statusLabels'])))},
    evolucaoLabels: {js_str(k['evoLabels'])},
    tentDia: {js_num(k['tentDia'])}, convDia: {js_num(k['convDia'])},
    showWpp: {show_wpp},
    wppTitle: 'WhatsApp — Smart Factory',
    wppDesc: 'Ações massivas exclusivas da campanha Smart Factory',
    wppKpiLabels: ['📤 Total Enviados','💬 Total Respostas','📊 Taxa de Resposta','🔇 Sem Resposta'],
    wppListLabels: ['Mensagens Enviadas','Informados','Solicitar E-mail','Sem Resposta'],
    wppPieLabels: ['Respostas','Sem Resposta'],
    wppEnv:'{k.get('wppEnv','-')}', wppResp:'{k.get('wppResp','-')}', wppTaxa:'{k.get('wppTaxa','-')}', wppSem:'{k.get('wppSem','-')}', wppInfo:'{k.get('wppInfo','-')}', wppEmail:'{k.get('wppEmail','-')}', wppPie:{wpp_pie}
  }},
  /* SMART_END */"""


def gerar_bloco_smart_agend(k):
    def js_str(lst): return '[' + ','.join(f"'{str(v).replace(chr(39), chr(92)+chr(39))}'" for v in lst) + ']'
    def js_num(lst): return '[' + ','.join(str(v) for v in lst) + ']'

    periodo = f"{k['evoLabels'][0]} — {k['evoLabels'][-1]}" if k['evoLabels'] else ''
    return f"""  /* SMART_AGEND_START */
  smart_agend: {{
    label: '— Smart Factory - Agendamentos', desc: 'Agendamentos da campanha Smart Factory — dados filtrados', periodo: '{periodo}',
    empresas: '{k['empresas']}', empresasLabel: '🏢 Empresas na Base',
    mediaLabel: '🔁 Média Tentativas/Empresa', mediaSub: 'por empresa',
    tentativas: '{k['tentativas']}', interessadosLabel: '📅 Agendamentos por Empresa', interessados: '{k['agendamentos']}', interessadosSub: 'Total de {k['agendamentosTotal']} agendamentos', conversao: '{k['conversao']}',
    decisor: '{k['decisor']}', decisorSub: 'Apenas Smart Factory - Agendamentos', media: '{k['media']}', trend: '',
    statusLabels: {js_str(k['statusLabels'])},
    statusData: {js_num(k['statusData'])}, statusColors:null,
    statusTooltips: {js_str(k.get('statusTooltips', ['']*len(k['statusLabels'])))},
    evolucaoLabels: {js_str(k['evoLabels'])},
    tentDia: {js_num(k['tentDia'])}, convDia: {js_num(k['convDia'])},
    showWpp: false,
    wppTitle: '', wppDesc: '',
    wppKpiLabels: [], wppListLabels: [], wppPieLabels: [],
    wppEnv:'-', wppResp:'-', wppTaxa:'-', wppSem:'-', wppInfo:'-', wppEmail:'-', wppPie:[0,1],
    showAgendTable: true,
    agendListaData: {js_str(k.get('agendListaData', []))},
    agendListaRazao: {js_str(k.get('agendListaRazao', []))},
    agendListaSol: {js_str(k.get('agendListaSol', []))}
  }},
  /* SMART_AGEND_END */"""

# ═══════════════════════════════════════════════════════════
# ATUALIZAÇÃO DO INDEX.HTML
# ═══════════════════════════════════════════════════════════

def atualizar_html(index_path, blocos):
    with open(index_path, 'r', encoding='utf-8') as f:
        conteudo = f.read()
    for marcador, novo_bloco in blocos.items():
        padrao = f'/\\* {marcador}_START \\*/.*?/\\* {marcador}_END \\*/'
        if not re.search(padrao, conteudo, re.DOTALL):
            raise ValueError(f'[ERRO] Marcadores {marcador} nao encontrados no index.html.')
        conteudo = re.sub(padrao, novo_bloco, conteudo, flags=re.DOTALL)
    with open(index_path, 'w', encoding='utf-8') as f:
        f.write(conteudo)

# ═══════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════

def main():
    print()
    print('=' * 50)
    print('  ATUALIZADOR — Smart Factory')
    print('=' * 50)

    os.chdir(os.path.dirname(os.path.abspath(__file__)))

    try:
        print('\n[0/4] Carregando de-para...')
        STATUS_MAP.update(ler_depara(DEPARA_PATH))

        print('\n[1/4] Calculando KPIs — Smart Factory...')
        ks = calcular_kpis_smart()

        print(f'  Empresas na Base  : {ks["empresas"]}')
        print(f'  Total Tentativas  : {ks["tentativas"]}')
        print(f'  Interessados      : {ks["interessados"]}')
        print(f'  Contatos Decisor  : {ks["decisor"]}')
        print(f'  Taxa Conversao    : {ks["conversao"]}')
        print(f'  Media Tent/Emp    : {ks["media"]}')
        print(f'  Status: {dict(zip(ks["statusLabels"][:5], ks["statusData"][:5]))} ...')
        if ks.get('showWpp'):
            print(f'  WhatsApp Enviados : {ks["wppEnv"]}')
            print(f'  WhatsApp Respostas: {ks["wppResp"]}  ({ks["wppTaxa"]})')
            print(f'  WhatsApp Sem Resp : {ks["wppSem"]}')
            print(f'  Informados        : {ks["wppInfo"]}  |  E-mail: {ks["wppEmail"]}')
        print('\n[2/3] Calculando KPIs — Smart Factory - Agendamentos...')
        try:
            ka = calcular_kpis_smart_agend()
            print(f'  Empresas na Base  : {ka["empresas"]}')
            print(f'  Total Tentativas  : {ka["tentativas"]}')
            print(f'  Contatos Decisor  : {ka["decisor"]}')
            print(f'  Agendamentos p/Empresa: {ka["agendamentos"]}  (total bruto: {ka["agendamentosTotal"]})')
            print(f'  Taxa Conversao    : {ka["conversao"]}')
            print(f'  Media Tent/Emp    : {ka["media"]}')
            print(f'  Status: {dict(zip(ka["statusLabels"][:5], ka["statusData"][:5]))} ...')
            print(f'  Lista Agendamentos: {len(ka.get("agendListaData", []))} linhas')
            bloco_agend = gerar_bloco_smart_agend(ka)
        except FileNotFoundError as e:
            print(f'  [AVISO] Smart Factory - Agendamentos nao encontrado. {e}')
            bloco_agend = None

        print('\n[3/3] Gerando blocos JavaScript e atualizando index.html...')
        blocos = {'SMART': gerar_bloco_smart(ks)}
        if bloco_agend:
            blocos['SMART_AGEND'] = bloco_agend
        atualizar_html(INDEX_HTML, blocos)

        print()
        print('=' * 50)
        print('  CONCLUIDO! index.html atualizado.')
        print('  Rode publicar.bat para enviar ao GitHub.')
        print('=' * 50)
        print()

    except Exception as e:
        print(f'\n[ERRO] {e}')
        import traceback; traceback.print_exc()

if __name__ == '__main__':
    main()
    input('Pressione ENTER para fechar...')
