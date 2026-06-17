# -*- coding: utf-8 -*-
"""
Atualizador Dashboard Firjan - Modulo URA
Le o arquivo BASE URA.xlsx e injeta linhas compactas no index.html.

Abas usadas:
  BASE DIGITAL   -> tipo Digital, data coluna B, classificacao coluna K
  Base Callflex  -> tipo Voz,     data coluna A, classificacao coluna AA

Cada linha valida conta como atendimento.
"""

import glob
import os
import re
import unicodedata
from datetime import datetime, timedelta, time

import openpyxl

PASTA = r'Arquivos\atualizaveis'
PREFIXO = 'BASE URA'
INDEX_HTML = r'index.html'

ABA_DIGITAL = 'BASE DIGITAL'
ABA_CALLFLEX = 'Base Callflex'

DIG_DATA = 1       # B - Data de criacao
DIG_CLASS = 10     # K - Classificacao Ura
DIG_TMA = 3        # D - Duracao em segundos
CALL_DATA = 0      # A - DATA
CALL_TMA = 19      # T - TMA
CALL_CLASS = 26    # AA - CLASSIFICACAO URA
CALL_MOTIVO = 28   # AC - Motivo_Tranferencia

CLASS_RETIDO = 0
CLASS_RESOLVIDO = 1
CLASS_ABANDONO = 2
CLASS_TRANSFERENCIA = 3
CLASS_OUTROS = 4


def encontrar_arquivo(pasta, prefixo):
    padrao = os.path.join(pasta, f'{prefixo}*.xlsx')
    arquivos = [a for a in glob.glob(padrao) if not os.path.basename(a).startswith('~$')]
    if not arquivos:
        raise FileNotFoundError(f'[ERRO] Arquivo nao encontrado: {prefixo}*.xlsx em {pasta}')
    return sorted(arquivos)[-1]


def to_dt(val):
    if isinstance(val, datetime):
        return val
    if isinstance(val, (int, float)) and val > 0:
        try:
            return datetime(1899, 12, 30) + timedelta(days=float(val))
        except Exception:
            return None
    s = str(val).strip() if val is not None else ''
    for fmt in ('%d/%m/%Y', '%d/%m/%Y %H:%M:%S', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d'):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    m = re.match(r'(\d{1,2})/(\d{1,2})/(\d{2,4})', s)
    if m:
        dd, mo, yy = m.groups()
        yy = int(yy) + (2000 if int(yy) < 100 else 0)
        return datetime(yy, int(mo), int(dd))
    return None


def to_ymd(val):
    d = to_dt(val)
    return d.year * 10000 + d.month * 100 + d.day if d else 0


def cel(row, idx):
    return row[idx] if len(row) > idx else None


def txt(v):
    return str(v).strip() if v is not None else ''


def norm(s):
    b = unicodedata.normalize('NFKD', txt(s).lower())
    return ''.join(c for c in b if not unicodedata.combining(c)).strip()


def class_code(v):
    n = norm(v)
    if n == 'transferencia humano':
        return CLASS_TRANSFERENCIA
    if n == 'resolvido na ura':
        return CLASS_RESOLVIDO
    if n == 'abandono':
        return CLASS_ABANDONO
    if n:
        return CLASS_OUTROS
    return CLASS_RETIDO


def to_seconds(v):
    if v is None:
        return 0
    if isinstance(v, timedelta):
        return int(round(v.total_seconds()))
    if isinstance(v, time):
        return v.hour * 3600 + v.minute * 60 + v.second
    if isinstance(v, datetime):
        return v.hour * 3600 + v.minute * 60 + v.second
    if isinstance(v, (int, float)):
        # Valores fracionarios do Excel representam parte de um dia; valores altos ja sao segundos.
        return int(round(float(v) * 86400)) if 0 < float(v) < 1 else int(round(float(v)))
    s = txt(v)
    if not s:
        return 0
    try:
        n = float(s.replace(',', '.'))
        return int(round(n * 86400)) if 0 < n < 1 else int(round(n))
    except ValueError:
        pass
    m = re.match(r'^(\d+):(\d{1,2})(?::(\d{1,2}))?$', s)
    if m:
        h, mi, se = m.groups()
        return int(h) * 3600 + int(mi) * 60 + int(se or 0)
    return 0


def motivo_valido(v):
    t = txt(v)
    n = norm(t)
    if not t or n in {'#n/d', '#n/a', 'n/d', 'na', 'nan'}:
        return ''
    return t


def ler_aba(ws, tipo_idx, col_data, col_class, col_tma=None, col_motivo=None, motivo_idx=None):
    rows = []
    total = 0
    motivo_idx = motivo_idx if motivo_idx is not None else {}
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if not any(c is not None and str(c).strip() for c in r):
            continue
        total += 1
        cls = class_code(cel(r, col_class))
        tma = to_seconds(cel(r, col_tma)) if col_tma is not None else 0
        mi = -1
        if col_motivo is not None and cls == CLASS_TRANSFERENCIA:
            mot = motivo_valido(cel(r, col_motivo))
            if mot:
                if mot not in motivo_idx:
                    motivo_idx[mot] = len(motivo_idx)
                mi = motivo_idx[mot]
        rows.append([to_ymd(cel(r, col_data)), tipo_idx, cls, tma, mi])
    return rows, total


def processar(caminho):
    print(f'  Lendo: {os.path.basename(caminho)}')
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    for aba in (ABA_DIGITAL, ABA_CALLFLEX):
        if aba not in wb.sheetnames:
            raise ValueError(f'Aba "{aba}" nao encontrada. Abas: {wb.sheetnames}')

    motivo_idx = {}
    rows_dig, total_dig = ler_aba(wb[ABA_DIGITAL], 0, DIG_DATA, DIG_CLASS, DIG_TMA)
    rows_voz, total_voz = ler_aba(wb[ABA_CALLFLEX], 1, CALL_DATA, CALL_CLASS, CALL_TMA, CALL_MOTIVO, motivo_idx)
    wb.close()

    rows = rows_dig + rows_voz
    motivos = [m for m, _ in sorted(motivo_idx.items(), key=lambda kv: kv[1])]
    print(f'  Digital: {total_dig} atendimentos | Voz: {total_voz} atendimentos | Total: {len(rows)}')
    print(f'  Motivos transferencia humano: {len(motivos)}')
    return ['Digital', 'Voz'], motivos, rows


def gerar_bloco(tipos, motivos, rows):
    def js_str(lst):
        return '[' + ','.join("'" + str(v).replace('\\', '\\\\').replace("'", "\\'") + "'" for v in lst) + ']'
    def js_rows(lst):
        return '[' + ','.join('[' + ','.join(str(c) for c in r) + ']' for r in lst) + ']'
    return (
        '/* URA_DATA_START */\n'
        f'const URA_TIPOS={js_str(tipos)};\n'
        f'const URA_MOTIVOS={js_str(motivos)};\n'
        f'const URA_ROWS={js_rows(rows)};\n'
        '/* URA_DATA_END */'
    )


def atualizar_html(index_path, bloco):
    with open(index_path, 'r', encoding='utf-8') as f:
        conteudo = f.read()
    padrao = r'/\* URA_DATA_START \*/.*?/\* URA_DATA_END \*/'
    if not re.search(padrao, conteudo, re.DOTALL):
        raise ValueError('[ERRO] Marcadores URA_DATA nao encontrados no index.html.')
    conteudo = re.sub(padrao, lambda m: bloco, conteudo, flags=re.DOTALL)
    with open(index_path, 'w', encoding='utf-8') as f:
        f.write(conteudo)


def main():
    print()
    print('=' * 50)
    print('  ATUALIZADOR - URA')
    print('=' * 50)
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    try:
        print('\n[1/2] Processando dados...')
        caminho = encontrar_arquivo(PASTA, PREFIXO)
        tipos, motivos, rows = processar(caminho)
        print('\n[2/2] Atualizando index.html...')
        atualizar_html(INDEX_HTML, gerar_bloco(tipos, motivos, rows))
        print('\n  CONCLUIDO! index.html atualizado.')
    except Exception as e:
        print(f'\n[ERRO] {e}')
        import traceback
        traceback.print_exc()


if __name__ == '__main__':
    main()
    try:
        input('Pressione ENTER para fechar...')
    except EOFError:
        pass
