# -*- coding: utf-8 -*-
"""
Atualizador Dashboard Firjan - Campanha Retomada da Trilha
Le o arquivo Mailing Excel e atualiza automaticamente o index.html

Como usar:
  - Coloque o arquivo Mailing mais recente em:
      Arquivos/nao_atualizaveis/Ativo_Retomada_da_Trilha/
  - De duplo clique em atualizar_e_publicar.bat
"""

import openpyxl
import re
import glob
import os
from collections import Counter, defaultdict
from datetime import datetime, timedelta

# ═══════════════════════════════════════════════════════════
# CONFIGURAÇÃO
# ═══════════════════════════════════════════════════════════

PASTA_MAILING = r'Arquivos\nao_atualizaveis\Ativo_Retomada_da_Trilha'
INDEX_HTML    = r'index.html'
DEPARA_PATH   = r'Arquivos\bases_apoio\tab_de-para.xlsx'

# STATUS_MAP é carregado dinamicamente do arquivo tab_de-para.xlsx
# (preenchido pela função ler_depara() no início da execução)
STATUS_MAP = {}

MES_PT = {1:'Jan',2:'Fev',3:'Mar',4:'Abr',5:'Mai',6:'Jun',
          7:'Jul',8:'Ago',9:'Set',10:'Out',11:'Nov',12:'Dez'}

# ═══════════════════════════════════════════════════════════
# FUNÇÕES AUXILIARES
# ═══════════════════════════════════════════════════════════

def fmt_num(n):
    """Número inteiro com separador de milhar PT-BR: 1.234"""
    return f"{n:,}".replace(",", ".")

def fmt_pct(n):
    """Percentual PT-BR: 86,05%"""
    return f"{n:.2f}%".replace(".", ",")

def fmt_dec(n):
    """Decimal PT-BR: 4,33"""
    return f"{n:.2f}".replace(".", ",")

def encontrar_arquivo(pasta, prefixo):
    """Encontra o arquivo Excel mais recente na pasta com o prefixo dado."""
    padrao = os.path.join(pasta, f'{prefixo}_*.xlsx')
    arquivos = [a for a in glob.glob(padrao) if not os.path.basename(a).startswith('~$')]
    if not arquivos:
        raise FileNotFoundError(
            f'\n[ERRO] Nenhum arquivo encontrado em:\n  {pasta}\n'
            f'Esperado: {prefixo}_XXXXXX.xlsx'
        )
    return sorted(arquivos)[-1]

def encontrar_mailing(pasta):
    return encontrar_arquivo(pasta, 'Mailing_Ativo_B+P_Retomada_da_Trilha')

def encontrar_retorno(pasta):
    return encontrar_arquivo(pasta, 'Retorno_Ativo_B+P_Retomada_da_Trilha')

def ler_depara(caminho):
    """Lê o arquivo tab_de-para.xlsx e retorna um dict {status_bruto: label_grafico}.
    Espera colunas: 'Status' (coluna A) e 'De-Para' (coluna B).
    Novos status adicionados no arquivo são capturados automaticamente."""
    mapa = {}
    try:
        wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
        ws = wb['status'] if 'status' in wb.sheetnames else wb.active  # aba "status"
        headers = None
        for row in ws.iter_rows(values_only=True):
            if headers is None:
                headers = [str(h).strip() if h else '' for h in row]
                continue
            if not any(row):
                continue
            status_bruto = row[0]
            label        = row[1]
            if status_bruto and label:
                mapa[str(status_bruto).strip()] = str(label).strip()
        wb.close()
        print(f'  De-para carregado: {len(mapa)} entradas de {os.path.basename(caminho)}')
    except FileNotFoundError:
        print(f'  [AVISO] Arquivo de-para nao encontrado: {caminho}')
        print(f'  Usando labels brutos do arquivo Mailing.')
    return mapa

def normalizar_status(raw):
    """Normaliza o status bruto para label do gráfico."""
    if raw is None:
        return None
    s = str(raw).strip().upper()
    for k, v in STATUS_MAP.items():
        if s == k.upper():
            return v
    return str(raw).strip()  # retorna como está se não mapeado

def to_datetime(val):
    """Converte valor de célula Excel para datetime.
    Trata tanto datetime objects quanto números seriais do Excel."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, (int, float)) and val > 0:
        # Número serial do Excel: dias desde 1899-12-30
        try:
            return datetime(1899, 12, 30) + timedelta(days=float(val))
        except Exception:
            return None
    return None

# ═══════════════════════════════════════════════════════════
# LEITURA E CÁLCULO
# ═══════════════════════════════════════════════════════════

def norm_tel(t):
    """Normaliza telefone removendo tudo que não é dígito."""
    return re.sub(r'\D', '', str(t)) if t else ''

# Decisor e Interessado determinados pelo LABEL do de-para:
# Decisor    = label não está em LABELS_NAO_DECISOR
# Interessado = label == 'Interessado'
LABELS_NAO_DECISOR = {'Telefonia', 'Tentativa', 'Engano', 'Alo'}
LABEL_INTERESSADO  = 'Interessado'


def _norm_status(s):
    import unicodedata
    b = unicodedata.normalize('NFKD', str(s).strip().lower())
    return ''.join(c for c in b if not unicodedata.combining(c))


# Grafico "Contatos de Sucesso": tabulacoes brutas (antes do de-para) consideradas sucesso.
# Variantes de grafia/acentuacao sao unificadas no mesmo rotulo canonico.
SUCESSO_LABEL_OVERRIDE = {
    'INTERESSADO(A)':      'Interessado',
    'Interessado':         'Interessado',
    'JÀ INFORMADO':        'Já Informado',
    'JÁ INFORMADO':        'Já Informado',
    'NÃO INTERESSADO(A)':  'Não Interessado',
    'Não Interessado':     'Não Interessado',
    'Retornar':            'Retornar',
}
SUCESSO_MAP = {_norm_status(k): v for k, v in SUCESSO_LABEL_OVERRIDE.items()}

# Grafico "Tentativas de Contato Sem Sucesso": tabulacoes sem interacao com o operador
# (variantes de grafia/acentuacao) sao agrupadas sob o rotulo unico "Tentativa"
SEM_OPERADOR_LABELS = ['Tel Não Atende / Ocupado', 'TEL NÃO ATENDE / OCUPADO',
                        'Fora de Area / Cx de Mensagens', 'FORA DE ÁREA / CX MENSAGENS',
                        'Ligação Muda']
SEM_OPERADOR_LABEL_CANON = 'Tentativa'
SEM_OPERADOR_NORM = {_norm_status(s) for s in SEM_OPERADOR_LABELS}

def calcular_mailing(caminho):
    """Lê o Mailing:
    - Empresas (CNPJs distintos)
    - Tentativas (datas preenchidas por tentativa 1-7)
    - Interessados (observações que começam com interessado/a)
    - Contatos com Decisor (Falamos com ≠ vazio e ≠ *)
    - Status (coluna Status via de-para)
    - Evolução diária (datas das tentativas)
    Retorna também conjuntos de telefones para deduplicação com Retorno.
    """
    print(f'  Lendo Mailing: {os.path.basename(caminho)}')
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb.active
    headers = None
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0: headers = [str(h) if h else '' for h in row]
        else: rows.append(row)
    wb.close()

    def idx(nome):
        for i, h in enumerate(headers):
            if nome.lower() in h.lower(): return i
        raise ValueError(f'Coluna nao encontrada: "{nome}"')

    cnpj_col   = idx('CNPJ')
    tel_col    = idx('Telefone')
    status_col = idx('Status')
    tent_data_cols, tent_falamos_cols, tent_obs_cols = [], [], []
    for n in range(1, 8):
        try:
            tent_data_cols.append(idx(f'Data e hora (tentativa {n})'))
            tent_falamos_cols.append(idx(f'Falamos com (tentativa {n})'))
            tent_obs_cols.append(idx(f'servação (tentativa {n})'))
        except ValueError: break

    cnpjs = set()
    total_tent = 0
    status_counter = Counter()
    raw_por_label  = defaultdict(set)
    naosucesso_counter = Counter()
    raw_por_label_ns   = defaultdict(set)
    decisor_count  = 0
    tel_interesse  = set()
    data_min = data_max = None

    for row in rows:
        if row[cnpj_col]: cnpjs.add(row[cnpj_col])
        tel = norm_tel(row[tel_col])

        st_raw   = row[status_col]
        raw      = str(st_raw).strip() if st_raw else ''
        raw_norm = _norm_status(raw) if raw else ''
        if raw_norm in SUCESSO_MAP:
            canon = SUCESSO_MAP[raw_norm]
            status_counter[canon] += 1
            raw_por_label[canon].add(raw)
            decisor_count += 1
        elif raw:
            ns_label = SEM_OPERADOR_LABEL_CANON if raw_norm in SEM_OPERADOR_NORM else raw
            naosucesso_counter[ns_label] += 1
            raw_por_label_ns[ns_label].add(raw)

        for di, fi, oi in zip(tent_data_cols, tent_falamos_cols, tent_obs_cols):
            data_val    = to_datetime(row[di])
            obs_val     = row[oi]

            if data_val:
                total_tent += 1
                if data_min is None or data_val < data_min:
                    data_min = data_val
                if data_max is None or data_val > data_max:
                    data_max = data_val
                if obs_val:
                    obs_str = str(obs_val).strip().lower()
                    if obs_str.startswith('interessado') or obs_str.startswith('interessada'):
                        tel_interesse.add(tel)

    return {
        '_empresas':      len(cnpjs),
        '_tentativas':    total_tent,
        '_statusCounter': status_counter,
        '_raw_por_label': dict(raw_por_label),
        '_naosucessoCounter': naosucesso_counter,
        '_raw_por_label_ns':  dict(raw_por_label_ns),
        '_decisorCount':  decisor_count,
        '_dataMin': data_min, '_dataMax': data_max,
        '_tel_interesse': tel_interesse,
    }


def calcular_retorno(caminho):
    """Lê o Retorno: cada linha = 1 tentativa.
    STATUS_NEGOCIO → de-para → label do gráfico.
    Retorna conjuntos de telefones para deduplicação com Mailing."""
    print(f'  Lendo Retorno: {os.path.basename(caminho)}')
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb.active
    headers = [str(h) if h else '' for h in next(ws.iter_rows(values_only=True))]
    wb_rows = list(ws.iter_rows(values_only=True, min_row=2))
    wb.close()

    data_idx = headers.index('DATA')
    sn_idx   = headers.index('STATUS_NEGOCIO')
    orig_idx = headers.index('ORIGEM')

    total_tent = 0
    status_counter = Counter()
    raw_por_label  = defaultdict(set)
    naosucesso_counter = Counter()
    raw_por_label_ns   = defaultdict(set)
    decisor_count = 0
    tel_interesse = set()
    data_min = data_max = None

    for row in wb_rows:
        total_tent += 1
        sn_raw   = row[sn_idx]
        sn_label = normalizar_status(sn_raw) if sn_raw else None
        tel      = norm_tel(row[orig_idx])
        raw      = str(sn_raw).strip() if sn_raw else ''
        raw_norm = _norm_status(raw) if raw else ''

        if raw_norm in SUCESSO_MAP:
            canon = SUCESSO_MAP[raw_norm]
            status_counter[canon] += 1
            raw_por_label[canon].add(raw)
            decisor_count += 1
        elif raw:
            ns_label = SEM_OPERADOR_LABEL_CANON if raw_norm in SEM_OPERADOR_NORM else raw
            naosucesso_counter[ns_label] += 1
            raw_por_label_ns[ns_label].add(raw)

        # Interessado = label == 'Interessado' (nao alterado por esta mudanca)
        if sn_label == LABEL_INTERESSADO:
            tel_interesse.add(tel)

        dt = row[data_idx]
        if not isinstance(dt, datetime): dt = to_datetime(dt)
        if dt:
            if data_min is None or dt < data_min:
                data_min = dt
            if data_max is None or dt > data_max:
                data_max = dt

    return {
        '_tentativas':    total_tent,
        '_statusCounter': status_counter,
        '_raw_por_label': dict(raw_por_label),
        '_naosucessoCounter': naosucesso_counter,
        '_raw_por_label_ns':  dict(raw_por_label_ns),
        '_decisorCount':  decisor_count,
        '_tel_interesse': tel_interesse,
        '_dataMin': data_min, '_dataMax': data_max,
    }


def combinar_kpis(m, r):
    """Soma Mailing + Retorno, deduplicando decisor e interessados por telefone."""
    empresas = m['_empresas']
    tent     = m['_tentativas'] + r['_tentativas']

    # Interessados: dedup por telefone (nao alterado por esta mudanca)
    tel_interesse_total = m['_tel_interesse'] | r['_tel_interesse']
    interesse = len(tel_interesse_total)

    # Decisor ("Contatos de Sucesso") = soma da quantidade de tabulacoes do grafico de sucesso
    decisor = m['_decisorCount'] + r['_decisorCount']

    taxa  = (interesse / decisor * 100) if decisor > 0 else 0
    media = tent / empresas if empresas > 0 else 0

    # Somar status (sucesso) dos dois arquivos
    status_total = m['_statusCounter'] + r['_statusCounter']
    naosucesso_total = m['_naosucessoCounter'] + r['_naosucessoCounter']

    # Mesclar raw_por_label dos dois arquivos
    raw_merged = defaultdict(set)
    for fonte in (m.get('_raw_por_label', {}), r.get('_raw_por_label', {})):
        for lbl, raws in fonte.items():
            raw_merged[lbl].update(raws)
    raw_merged_ns = defaultdict(set)
    for fonte in (m.get('_raw_por_label_ns', {}), r.get('_raw_por_label_ns', {})):
        for lbl, raws in fonte.items():
            raw_merged_ns[lbl].update(raws)

    datas = [d for d in (m.get('_dataMin'), m.get('_dataMax'), r.get('_dataMin'), r.get('_dataMax')) if d]
    data_min, data_max = (min(datas), max(datas)) if datas else (None, None)
    if data_min and data_max:
        periodo = f"{data_min.day:02d}/{MES_PT[data_min.month]} — {data_max.day:02d}/{MES_PT[data_max.month]}"
    else:
        periodo = ''

    st_items = status_total.most_common()
    st_tooltips = [', '.join(sorted(raw_merged.get(s[0], set()))) for s in st_items]
    ns_items = naosucesso_total.most_common()
    ns_tooltips = [', '.join(sorted(raw_merged_ns.get(s[0], set()))) for s in ns_items]

    return {
        # Brutos para somar campanhas (todas)
        '_empresas':      empresas,
        '_tentativas':    tent,
        '_interessados':  interesse,
        '_decisor':       decisor,
        '_statusCounter':  status_total,
        '_raw_por_label':  dict(raw_merged),
        '_evolucao':       {},
        # Formatados para o dashboard
        'empresas':      fmt_num(empresas),
        'tentativas':    fmt_num(tent),
        'interessados':  fmt_num(interesse),
        'conversao':     fmt_pct(taxa),
        'decisor':       fmt_num(decisor),
        'media':         fmt_dec(media),
        'periodo':       periodo,
        'statusLabels':  [s[0] for s in st_items],
        'statusData':    [s[1] for s in st_items],
        'statusTooltips': st_tooltips,
        'naosucessoLabels':   [s[0] for s in ns_items],
        'naosucessoData':     [s[1] for s in ns_items],
        'naosucessoTooltips': ns_tooltips,
        'evoLabels':     [], 'tentDia': [], 'convDia': [],
    }


def calcular_kpis(pasta):
    """Lê Mailing + Retorno e combina os KPIs."""
    m = calcular_mailing(encontrar_mailing(pasta))
    r = calcular_retorno(encontrar_retorno(pasta))
    return combinar_kpis(m, r)

# ═══════════════════════════════════════════════════════════
# SOMA DE CAMPANHAS → TODAS
# ═══════════════════════════════════════════════════════════

def somar_campanhas(lista_kpis):
    """Recebe uma lista de kpis (dicts com campos _brutos) e retorna
    um dict formatado com a soma de todas as campanhas."""
    total_emp   = sum(k['_empresas']     for k in lista_kpis)
    total_tent  = sum(k['_tentativas']   for k in lista_kpis)
    total_int   = sum(k['_interessados'] for k in lista_kpis)
    total_dec   = sum(k['_decisor']      for k in lista_kpis)

    taxa  = (total_int / total_dec * 100) if total_dec > 0 else 0
    media = total_tent / total_emp if total_emp > 0 else 0

    # Somar contadores de status
    status_total = Counter()
    for k in lista_kpis:
        status_total += k['_statusCounter']
    st_items  = status_total.most_common()
    st_labels = [s[0] for s in st_items]
    st_data   = [s[1] for s in st_items]

    # Mesclar status brutos por label (uniao entre campanhas) p/ a legenda
    raw_merged = defaultdict(set)
    for k in lista_kpis:
        for lbl, raws in k.get('_raw_por_label', {}).items():
            raw_merged[lbl] |= set(raws)
    st_tooltips = [', '.join(sorted(raw_merged.get(lbl, set()))) for lbl in st_labels]

    # Mesclar evolução diária (somar por dia)
    evo_total = {}
    for k in lista_kpis:
        for dk, vals in k['_evolucao'].items():
            if dk not in evo_total:
                evo_total[dk] = {'date': vals['date'], 'tent': 0, 'conv': 0}
            evo_total[dk]['tent'] += vals['tent']
            evo_total[dk]['conv'] += vals['conv']

    dias_ord   = sorted(evo_total.items(), key=lambda x: x[1]['date'] or datetime(9999,12,31).date())
    evo_labels = [d[0] for d in dias_ord]
    tent_dia   = [d[1]['tent'] for d in dias_ord]
    conv_dia   = [d[1]['conv'] for d in dias_ord]

    return {
        'empresas':     fmt_num(total_emp),
        'tentativas':   fmt_num(total_tent),
        'interessados': fmt_num(total_int),
        'conversao':    fmt_pct(taxa),
        'decisor':      fmt_num(total_dec),
        'media':        fmt_dec(media),
        'statusLabels': st_labels,
        'statusData':   st_data,
        'statusTooltips': st_tooltips,
        'evoLabels':    evo_labels,
        'tentDia':      tent_dia,
        'convDia':      conv_dia,
    }

# ═══════════════════════════════════════════════════════════
# GERAÇÃO DO BLOCO JS
# ═══════════════════════════════════════════════════════════

def gerar_bloco_retomada(k):
    """Monta o bloco JavaScript da campanha Retomada da Trilha."""

    def js_str_array(lst):
        return '[' + ','.join(f"'{str(v).replace(chr(39), chr(92)+chr(39))}'" for v in lst) + ']'

    def js_num_array(lst):
        return '[' + ','.join(str(v) for v in lst) + ']'

    periodo = k.get('periodo', '')
    bloco = f"""  /* RETOMADA_START */
  retomada: {{
    label: '— Retomada da Trilha', desc: 'Campanha Retomada da Trilha — dados filtrados', periodo: '{periodo}',
    empresas: '{k['empresas']}', empresasLabel: '🏢 Empresas na Base',
    mediaLabel: '🔁 Média Tentativas/Empresa', mediaSub: 'por empresa',
    tentativas: '{k['tentativas']}', interessados: '{k['interessados']}', conversao: '{k['conversao']}',
    decisor: '{k['decisor']}', decisorLabel: '👤 Contatos de Sucesso', decisorSub: 'Apenas Retomada da Trilha', media: '{k['media']}', trend: '',
    distTitle: 'Contatos de Sucesso',
    statusLabels: {js_str_array(k['statusLabels'])},
    statusData: {js_num_array(k['statusData'])}, statusColors:null,
    statusTooltips: {js_str_array(k.get('statusTooltips', ['']*len(k['statusLabels'])))},
    evoTitle: 'Tentativas de Contato Sem Sucesso', evoBar: true,
    naosucessoLabels: {js_str_array(k.get('naosucessoLabels', []))},
    naosucessoData: {js_num_array(k.get('naosucessoData', []))},
    naosucessoTooltips: {js_str_array(k.get('naosucessoTooltips', []))},
    evolucaoLabels: [], tentDia: [], convDia: [],
    showWpp: false,
    wppTitle: 'WhatsApp', wppDesc: '',
    wppKpiLabels: ['📤 Total Enviados','💬 Total Respostas','📊 Taxa de Resposta','🔇 Sem Resposta'],
    wppListLabels: ['Mensagens Enviadas','Informados','Solicitar E-mail','Sem Resposta'],
    wppPieLabels: ['Respostas','Sem Resposta'],
    wppEnv:'-', wppResp:'-', wppTaxa:'-', wppSem:'-', wppInfo:'-', wppEmail:'-', wppPie:[0,1]
  }},
  /* RETOMADA_END */"""
    return bloco

def gerar_bloco_todas(k):
    """Monta o bloco JavaScript da campanha Todas (soma de todas)."""
    def js_str_array(lst):
        return '[' + ','.join(f"'{str(v).replace(chr(39), chr(92)+chr(39))}'" for v in lst) + ']'
    def js_num_array(lst):
        return '[' + ','.join(str(v) for v in lst) + ']'

    periodo = f"{k['evoLabels'][0]} — {k['evoLabels'][-1]}" if k['evoLabels'] else ''
    return f"""  /* TODAS_START */
  todas: {{
    label: '', desc: 'Todas as campanhas — Retomada da Trilha + Smart Factory + Cursos Técnicos Niterói', periodo: '',
    empresas: '{k['empresas']}', empresasLabel: '🏢 Empresas / Inscrições',
    mediaLabel: '🔁 Média Tentativas/Empresa', mediaSub: 'por empresa/inscrição',
    tentativas: '{k['tentativas']}', interessados: '{k['interessados']}', conversao: '{k['conversao']}',
    decisor: '{k['decisor']}', decisorSub: 'Retomada + Smart Factory + Niterói', media: '{k['media']}', trend: '',
    statusLabels: {js_str_array(k['statusLabels'])},
    statusData: {js_num_array(k['statusData'])}, statusColors:null,
    statusTooltips: {js_str_array(k.get('statusTooltips', ['']*len(k['statusLabels'])))},
    evolucaoLabels: {js_str_array(k['evoLabels'])},
    tentDia: {js_num_array(k['tentDia'])}, convDia: {js_num_array(k['convDia'])},
    showWpp: false,
    wppTitle: 'WhatsApp', wppDesc: '',
    wppKpiLabels: ['📤 Total Enviados','💬 Total Respostas','📊 Taxa de Resposta','🔇 Sem Resposta'],
    wppListLabels: ['Mensagens Enviadas','Informados','Solicitar E-mail','Sem Resposta'],
    wppPieLabels: ['Respostas','Sem Resposta'],
    wppEnv:'0', wppResp:'0', wppTaxa:'0,00%', wppSem:'0', wppInfo:'0', wppEmail:'0', wppPie:[0,1]
  }},
  /* TODAS_END */"""

# ═══════════════════════════════════════════════════════════
# ATUALIZAÇÃO DO INDEX.HTML
# ═══════════════════════════════════════════════════════════

def atualizar_html(index_path, blocos):
    """blocos: dict com chave=marcador, valor=novo bloco JS."""
    with open(index_path, 'r', encoding='utf-8') as f:
        conteudo = f.read()

    for marcador, novo_bloco in blocos.items():
        inicio = f'/\\* {marcador}_START \\*/'
        fim    = f'/\\* {marcador}_END \\*/'
        padrao = f'{inicio}.*?{fim}'
        if not re.search(padrao, conteudo, re.DOTALL):
            raise ValueError(f'\n[ERRO] Marcadores /* {marcador}_START */ e /* {marcador}_END */ não encontrados.\n')
        conteudo = re.sub(padrao, novo_bloco, conteudo, flags=re.DOTALL)

    with open(index_path, 'w', encoding='utf-8') as f:
        f.write(conteudo)

# ═══════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════

def main():
    print()
    print('=' * 50)
    print('  ATUALIZADOR — Retomada da Trilha')
    print('=' * 50)

    # Muda para a pasta do script (garante que os caminhos relativos funcionam)
    os.chdir(os.path.dirname(os.path.abspath(__file__)))

    try:
        # 0. Carregar de-para do arquivo externo
        print('\n[0/3] Carregando de-para...')
        STATUS_MAP.update(ler_depara(DEPARA_PATH))

        # 1. Ler campanha Retomada da Trilha (Mailing + Retorno)
        print('\n[1/3] Calculando KPIs — Retomada da Trilha...')
        kpis_retomada = calcular_kpis(PASTA_MAILING)

        print(f'  Empresas na Base : {kpis_retomada["empresas"]}')
        print(f'  Total Tentativas : {kpis_retomada["tentativas"]}')
        print(f'  Interessados     : {kpis_retomada["interessados"]}')
        print(f'  Contatos Decisor : {kpis_retomada["decisor"]}')
        print(f'  Taxa Conversão   : {kpis_retomada["conversao"]}')
        print(f'  Média Tent/Emp   : {kpis_retomada["media"]}')

        # 2. Somar todas as campanhas disponíveis
        print('\n[2/3] Calculando soma (Todas as Campanhas)...')
        todas_kpis_list = [kpis_retomada]
        # Quando Smart Factory for adicionado, incluir aqui:
        # if kpis_smart: todas_kpis_list.append(kpis_smart)
        kpis_todas = somar_campanhas(todas_kpis_list)
        print(f'  Empresas (soma)  : {kpis_todas["empresas"]}')
        print(f'  Tentativas (soma): {kpis_todas["tentativas"]}')
        print(f'  Interessados (s) : {kpis_todas["interessados"]}')

        # 3. Atualizar HTML com todos os blocos
        print('\n[3/3] Atualizando index.html...')
        blocos = {
            'TODAS':    gerar_bloco_todas(kpis_todas),
            'RETOMADA': gerar_bloco_retomada(kpis_retomada),
        }
        atualizar_html(INDEX_HTML, blocos)

        print()
        print('=' * 50)
        print('  CONCLUÍDO! index.html atualizado.')
        print('  Agora rode o publicar.bat para enviar ao GitHub.')
        print('=' * 50)
        print()

    except FileNotFoundError as e:
        print(e)
    except Exception as e:
        print(f'\n[ERRO INESPERADO] {e}')
        import traceback; traceback.print_exc()

if __name__ == '__main__':
    main()
    input('Pressione ENTER para fechar...')
