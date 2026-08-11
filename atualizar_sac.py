# -*- coding: utf-8 -*-
"""
Atualizador Dashboard Firjan - Modulo SAC
Le a planilha SAC do SharePoint (link publico/anonimo) e injeta os dados brutos
no index.html. Os filtros (Data, Regional, Tipo de Registro, Canal) e os KPIs
sao recalculados no navegador.

Aba: "Sac"
Colunas usadas (0-based):
  A(0)  = Protocolo         -> Total de SACs (linhas com este campo preenchido)
  C(2)  = Data do Atendimento -> filtro Data + base do Tempo Medio
  E(4)  = Canal             -> filtro
  I(8)  = Regional          -> filtro
  J(9)  = Tipo de Registro  -> filtro
  M(12) = Data de Finalizacao -> Tempo Medio = media(M - C)
  N(13) = Nivel de Satisfacao -> Satisfeito / Insatisfeito
"""

import urllib.request
import http.cookiejar
import io
import re
import os
from datetime import datetime, timedelta

# ═══════════════════════════════════════════════════════════
# CONFIGURAÇÃO
# ═══════════════════════════════════════════════════════════

# Link de compartilhamento do SharePoint (somente leitura/consulta).
# O ?download=1 faz o SharePoint devolver o arquivo .xlsx direto.
SAC_URL = ('https://ddmadvbr-my.sharepoint.com/:x:/g/personal/'
           'fernanda_castro_grupoddm_com_br/'
           'IQCobMqlyAWfT7aBxT5EGURCAaAxahCZQjDD0eCyvVChMv0?download=1')

# Link "limpo" (sem ?download=1) para registrar no arquivo de erros
SAC_LINK = ('https://ddmadvbr-my.sharepoint.com/:x:/g/personal/'
            'fernanda_castro_grupoddm_com_br/'
            'IQCobMqlyAWfT7aBxT5EGURCAaAxahCZQjDD0eCyvVChMv0')

ABA_SAC    = 'Sac'
INDEX_HTML = r'index.html'
ERROS_TXT  = r'erros_sac.txt'

# Tempo de resposta acima deste limite (ou negativo) e tratado como outlier
# e EXCLUIDO do calculo do Tempo Medio (registrado em erros_sac.txt).
MAX_DELTA_DIAS = 30

# --- Colunas fonte 1: SharePoint ---
COL_PROT    = 0   # A  Protocolo
COL_DATA    = 2   # C  Data do Atendimento
COL_CANAL   = 4   # E  Canal
COL_PROD    = 5   # F  Produto/Servico
COL_UNI     = 7   # H  Unidade
COL_ENT     = 8   # I  Entidade
COL_REG     = 9   # J  Regional
COL_TIPO    = 10  # K  Tipo de Registro
COL_FIM     = 13  # N  Data de Finalizacao
COL_SAT     = 14  # O  Nivel de Satisfacao
COL_ASSUNTO = 19  # T  Assunto
COL_IMPACTO = 21  # V  Impacto (so existe na Fonte 1 / SharePoint)
COL_ENCAM   = 3   # D  Data de Encaminhamento (bloco "Impactos por Unidade")
COL_ASSDET  = 20  # U  Detalhe do Assunto (bloco "Impactos por Unidade")

# --- Colunas fonte 2: Google Sheets ---
SAC2_URL = ('https://docs.google.com/spreadsheets/d/'
            '1z-4hnVB7JRoqTZZ0yVrZ0KONKR3RlualzrrfM3vw9xs/'
            'export?format=xlsx&gid=1326039238')

C2_PROT    = 0   # A  Protocolo
C2_DATA    = 2   # C  Data do Atendimento
C2_CANAL   = 4   # E  Canal
C2_PROD    = 5   # F  Produto/Servico
C2_TIPO    = 8   # I  Tipo de Registro
C2_FIM     = 11  # L  Data de Finalizacao
C2_SAT     = 12  # M  Nivel de Satisfacao
C2_ASSUNTO = 17  # R  Assunto
C2_UNI     = 21  # V  Unidade
C2_ENT     = 22  # W  Entidade
C2_REG     = 23  # X  Regional
C2_ENCAM   = 3   # D  Data de Encaminhamento (bloco "Impactos por Unidade")
C2_ASSDET  = 18  # S  Detalhe do Assunto (bloco "Impactos por Unidade"; Fonte 2 nao tem Impacto)

# ═══════════════════════════════════════════════════════════
# FUNÇÕES
# ═══════════════════════════════════════════════════════════

def baixar_xlsx(url):
    """Baixa o .xlsx do SharePoint usando cookie jar (link anonimo)."""
    print('  Baixando planilha SAC do SharePoint...')
    cj = http.cookiejar.CookieJar()
    op = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(cj))
    op.addheaders = [('User-Agent', 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)')]
    resp = op.open(url, timeout=60)
    data = resp.read()
    ctype = resp.headers.get('Content-Type', '')
    if 'spreadsheet' not in ctype and 'octet-stream' not in ctype:
        raise RuntimeError(f'Resposta nao e xlsx (Content-Type: {ctype}). '
                           f'O link pode ter mudado de permissao.')
    print(f'  OK ({len(data):,} bytes)')
    return data


def baixar_gsheets(url):
    """Baixa o .xlsx exportado do Google Sheets (link publico)."""
    print('  Baixando planilha SAC complementar do Google Sheets...')
    req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
    resp = urllib.request.urlopen(req, timeout=60)
    data = resp.read()
    print(f'  OK ({len(data):,} bytes)')
    return data


_HOJE = datetime.now()

def to_dt(val):
    d = None
    if isinstance(val, datetime):
        d = val
    elif isinstance(val, (int, float)) and val > 0:
        try:
            d = datetime(1899, 12, 30) + timedelta(days=float(val))
        except Exception:
            d = None
    elif isinstance(val, str):
        m = re.match(r'(\d{1,2})/(\d{1,2})/(\d{2,4})', val.strip())
        if m:
            dd, mo, yy = m.groups()
            yy = int(yy); yy = yy + 2000 if yy < 100 else yy
            try:
                d = datetime(yy, int(mo), int(dd))
            except ValueError:
                d = None
    if d is None:
        return None
    # Correcao BR x EUA: se a data caiu no FUTURO, troca dia<->mes
    if d > _HOJE:
        try:
            sw = d.replace(month=d.day, day=d.month)
            if sw <= _HOJE:
                d = sw
        except ValueError:
            pass
    return d


def cel(row, idx):
    if len(row) <= idx:
        return None
    return row[idx]


def sat_code(v):
    """1=Satisfeito, 2=Insatisfeito, 3=Nao Avaliado, 0=outros."""
    s = str(v).strip().lower() if v is not None else ''
    if s == 'satisfeito':
        return 1
    if s == 'insatisfeito':
        return 2
    if s in ('não avaliado', 'nao avaliado'):
        return 3
    return 0


def fmt_raw(v):
    if v is None:
        return ''
    if isinstance(v, datetime):
        return v.strftime('%d/%m/%Y %H:%M') if (v.hour or v.minute) else v.strftime('%d/%m/%Y')
    return str(v).strip()


def normalizar_unidade(v):
    """Tudo que nao comeca com 'Unidade' (inclusive em branco) vira 'Firjan (Outras Áreas)'."""
    s = str(v).strip() if v is not None else ''
    if s.lower().startswith('unidade'):
        return s
    return 'Firjan (Outras Áreas)'


def processar(xlsx_bytes):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    if ABA_SAC not in wb.sheetnames:
        raise ValueError(f'Aba "{ABA_SAC}" nao encontrada. Abas: {wb.sheetnames}')
    ws = wb[ABA_SAC]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()
    headers1 = [fmt_raw(v) for v in all_rows[0]] if all_rows else []
    rows = all_rows[1:]
    raw_rows1 = []

    canal_list, reg_list, tipo_list = [], [], []
    canal_idx, reg_idx, tipo_idx = {}, {}, {}
    ent_list, assunto_list, prod_list, uni_list = [], [], [], []
    ent_idx, assunto_idx, prod_idx, uni_idx = {}, {}, {}, {}
    impacto_list, impacto_idx = [], {}
    imp_uni_list, imp_uni_idx = [], {}
    imp_assunto_list, imp_assunto_idx = [], {}
    imp_nivel_list, imp_nivel_idx = [], {}

    def get_idx(val, lst, mp):
        v = str(val).strip() if val is not None else ''
        if v not in mp:
            mp[v] = len(lst)
            lst.append(v)
        return mp[v]

    def get_idx_nf(val, lst, mp):
        v = str(val).strip() if (val is not None and str(val).strip()) else 'Não Informado'
        if v not in mp:
            mp[v] = len(lst)
            lst.append(v)
        return mp[v]

    data_rows = []
    imp_rows = []
    erros = []
    outros_set = set()
    total = sat = insat = 0
    soma_delta = 0
    n_delta = 0
    limite_seg = MAX_DELTA_DIAS * 86400

    for i, r in enumerate(rows):
        linha_excel = i + 2  # +2: header (linha 1) + base 0
        prot = cel(r, COL_PROT)
        if prot in (None, ''):
            continue  # Total de SACs = linhas com Protocolo (col A)
        total += 1

        c = to_dt(cel(r, COL_DATA))
        m = to_dt(cel(r, COL_FIM))
        dt = (c.year * 10000 + c.month * 100 + c.day) if c else 0

        ci = get_idx(cel(r, COL_CANAL), canal_list, canal_idx)
        ri = get_idx(cel(r, COL_REG),   reg_list,   reg_idx)
        ti = get_idx(cel(r, COL_TIPO),  tipo_list,  tipo_idx)
        ei = get_idx(cel(r, COL_ENT), ent_list, ent_idx)   # Entidade (col I, direta)
        ui = get_idx(normalizar_unidade(cel(r, COL_UNI)), uni_list, uni_idx)   # Unidade (col H) -> ranking
        aci = get_idx(cel(r, COL_ASSUNTO), assunto_list, assunto_idx)
        pi = get_idx(cel(r, COL_PROD),    prod_list,    prod_idx)
        ii = get_idx(cel(r, COL_IMPACTO), impacto_list, impacto_idx)
        sc = sat_code(cel(r, COL_SAT))
        if sc == 1:
            sat += 1
        elif sc == 2:
            insat += 1
        elif sc == 0:  # "Outros" — guarda os rotulos para a legenda
            raw = cel(r, COL_SAT)
            outros_set.add('(em branco)' if raw in (None, '') else str(raw).strip())

        dl = None
        if c and m:
            d = int((m - c).total_seconds())
            motivo = None
            if d < 0:
                motivo = 'Data de Finalizacao (M) ANTERIOR a Data do Atendimento (C)'
            elif d > limite_seg:
                motivo = f'Tempo de resposta acima do limite de {MAX_DELTA_DIAS} dias'
            if motivo:
                erros.append({
                    'linha': linha_excel,
                    'protocolo': str(prot).strip(),
                    'atendimento': c.strftime('%d/%m/%Y %H:%M'),
                    'finalizacao': m.strftime('%d/%m/%Y %H:%M'),
                    'delta_dias': round(d / 86400, 2),
                    'motivo': motivo,
                })
            else:
                soma_delta += d
                n_delta += 1
                dl = d

        data_rows.append([dt, ci, ri, ti, sc, dl, ei, aci, pi, ui, ii])
        raw_rows1.append([fmt_raw(v) for v in r])

        imp_raw = cel(r, COL_IMPACTO)
        if imp_raw is not None and str(imp_raw).strip():  # sem Impacto = fora do bloco (nao existe "Nao Informado")
            enc = to_dt(cel(r, COL_ENCAM))
            dtEnc = (enc.year * 10000 + enc.month * 100 + enc.day) if enc else 0
            iu = get_idx(normalizar_unidade(cel(r, COL_UNI)), imp_uni_list, imp_uni_idx)
            ia = get_idx_nf(cel(r, COL_ASSDET), imp_assunto_list, imp_assunto_idx)
            iv = get_idx(str(imp_raw).strip(), imp_nivel_list, imp_nivel_idx)
            imp_rows.append([dtEnc, iu, ia, iv])

    tmr = (soma_delta / n_delta) if n_delta else 0
    print(f'  Total de SACs        : {total}')
    print(f'  Satisfeitos          : {sat}')
    print(f'  Insatisfeitos        : {insat}')
    print(f'  Tempo medio (validos): {tmr/3600:.1f}h = {tmr/86400:.2f} dias '
          f'({n_delta} validos | {len(erros)} outliers ignorados)')
    print(f'  Canais: {len(canal_list)} | Regionais: {len(reg_list)} | Tipos: {len(tipo_list)}')
    outros_labels = sorted(outros_set, key=lambda s: (s != '(em branco)', s.lower()))
    print(f'  "Outros" abrange: {outros_labels}')
    print(f'  Entidades: {len([e for e in ent_list if e])} | Assuntos: {len([a for a in assunto_list if a])} | Produtos: {len([p for p in prod_list if p])}')
    print(f'  Niveis de Impacto: {[i for i in impacto_list if i]}')
    print(f'  Impactos por Unidade: {len(imp_rows)} linhas | Unidades: {len(imp_uni_list)} | '
          f'Detalhes de Assunto: {len(imp_assunto_list)} | Niveis: {imp_nivel_list}')

    extras = {
        'ent': ent_list, 'assunto': assunto_list, 'prod': prod_list, 'uni': uni_list, 'impacto': impacto_list,
        'impUni': imp_uni_list, 'impUniIdx': imp_uni_idx,
        'impAssunto': imp_assunto_list, 'impAssuntoIdx': imp_assunto_idx,
        'impNivel': imp_nivel_list, 'impNivelIdx': imp_nivel_idx,
    }
    return canal_list, reg_list, tipo_list, data_rows, erros, outros_labels, extras, headers1, raw_rows1, imp_rows


def processar2(xlsx_bytes, canal_list, canal_idx, reg_list, reg_idx, tipo_list, tipo_idx,
               ent_list, ent_idx, assunto_list, assunto_idx, prod_list, prod_idx,
               uni_list, uni_idx, impacto_list, impacto_idx,
               imp_uni_list, imp_uni_idx, imp_assunto_list, imp_assunto_idx,
               imp_nivel_list, imp_nivel_idx):
    """Processa a fonte 2 (Google Sheets) reaproveitando as listas da fonte 1."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb.active
    all_rows2 = list(ws.iter_rows(values_only=True))
    wb.close()
    headers2 = [fmt_raw(v) for v in all_rows2[0]] if all_rows2 else []
    rows = all_rows2[1:]
    raw_rows2 = []

    def get_idx(val, lst, mp):
        v = str(val).strip() if val is not None else ''
        if v not in mp:
            mp[v] = len(lst)
            lst.append(v)
        return mp[v]

    def get_idx_nf(val, lst, mp):
        v = str(val).strip() if (val is not None and str(val).strip()) else 'Não Informado'
        if v not in mp:
            mp[v] = len(lst)
            lst.append(v)
        return mp[v]

    data_rows = []
    imp_rows = []
    erros = []
    outros_set = set()
    total = sat = insat = 0
    soma_delta = 0
    n_delta = 0
    limite_seg = MAX_DELTA_DIAS * 86400

    for i, r in enumerate(rows):
        linha_excel = i + 2
        prot = cel(r, C2_PROT)
        if prot in (None, ''):
            continue
        total += 1

        c = to_dt(cel(r, C2_DATA))
        m = to_dt(cel(r, C2_FIM))
        dt = (c.year * 10000 + c.month * 100 + c.day) if c else 0

        ci  = get_idx(cel(r, C2_CANAL),   canal_list,   canal_idx)
        ri  = get_idx(cel(r, C2_REG),     reg_list,     reg_idx)
        ti  = get_idx(cel(r, C2_TIPO),    tipo_list,    tipo_idx)
        ei  = get_idx(cel(r, C2_ENT),     ent_list,     ent_idx)
        ui  = get_idx(normalizar_unidade(cel(r, C2_UNI)), uni_list, uni_idx)
        aci = get_idx(cel(r, C2_ASSUNTO), assunto_list, assunto_idx)
        pi  = get_idx(cel(r, C2_PROD),    prod_list,    prod_idx)
        ii  = get_idx(None, impacto_list, impacto_idx)  # Fonte 2 nao tem coluna Impacto
        sc  = sat_code(cel(r, C2_SAT))
        if sc == 1:
            sat += 1
        elif sc == 2:
            insat += 1
        elif sc == 0:
            raw = cel(r, C2_SAT)
            outros_set.add('(em branco)' if raw in (None, '') else str(raw).strip())

        dl = None
        if c and m:
            d = int((m - c).total_seconds())
            motivo = None
            if d < 0:
                motivo = 'Data de Finalizacao (L) ANTERIOR a Data do Atendimento (C)'
            elif d > limite_seg:
                motivo = f'Tempo de resposta acima do limite de {MAX_DELTA_DIAS} dias'
            if motivo:
                erros.append({
                    'linha': linha_excel,
                    'protocolo': str(prot).strip(),
                    'atendimento': c.strftime('%d/%m/%Y %H:%M'),
                    'finalizacao': m.strftime('%d/%m/%Y %H:%M'),
                    'delta_dias': round(d / 86400, 2),
                    'motivo': motivo,
                })
            else:
                soma_delta += d
                n_delta += 1
                dl = d

        data_rows.append([dt, ci, ri, ti, sc, dl, ei, aci, pi, ui, ii])
        raw_rows2.append([fmt_raw(v) for v in r])

        # Fonte 2 nao tem coluna Impacto -> nao entra no bloco "Impactos por Unidade"

    print(f'  [Fonte 2] Total: {total} | Satisfeitos: {sat} | Insatisfeitos: {insat} | Outliers: {len(erros)}')
    return data_rows, erros, headers2, raw_rows2, imp_rows


def gerar_bloco(canal_list, reg_list, tipo_list, data_rows, outros_labels=None, extras=None,
                headers1=None, raw_rows1=None, headers2=None, raw_rows2=None, imp_rows=None):
    def js_str(lst):
        return '[' + ','.join("'" + str(v).replace('\\', '\\\\').replace("'", "\\'") + "'" for v in lst) + ']'
    def js_rows(rows):
        partes = []
        for r in rows:
            cells = []
            for c in r:
                cells.append('null' if c is None else str(c))
            partes.append('[' + ','.join(cells) + ']')
        return '[' + ','.join(partes) + ']'
    def js_raw(rows):
        partes = []
        for r in rows:
            cells = ["'" + str(v).replace('\\', '\\\\').replace('\r', '').replace('\n', ' ').replace("'", "\\'") + "'" for v in r]
            partes.append('[' + ','.join(cells) + ']')
        return '[' + ','.join(partes) + ']'

    extras = extras or {}
    return (
        '/* SAC_DATA_START */\n'
        f'const SAC_CANAL={js_str(canal_list)};\n'
        f'const SAC_REG={js_str(reg_list)};\n'
        f'const SAC_TIPO={js_str(tipo_list)};\n'
        f'const SAC_ENT={js_str(extras.get("ent", []))};\n'
        f'const SAC_ASSUNTO={js_str(extras.get("assunto", []))};\n'
        f'const SAC_PROD={js_str(extras.get("prod", []))};\n'
        f'const SAC_UNI={js_str(extras.get("uni", []))};\n'
        f'const SAC_IMPACTO={js_str(extras.get("impacto", []))};\n'
        f'const SAC_OUTROS={js_str(outros_labels or [])};\n'
        f'const SAC_ROWS={js_rows(data_rows)};\n'
        f'const SAC_HEADERS1={js_str(headers1 or [])};\n'
        f'const SAC_RAW1={js_raw(raw_rows1 or [])};\n'
        f'const SAC_HEADERS2={js_str(headers2 or [])};\n'
        f'const SAC_RAW2={js_raw(raw_rows2 or [])};\n'
        f'const SAC_IMP_UNI={js_str(extras.get("impUni", []))};\n'
        f'const SAC_IMP_ASSUNTO={js_str(extras.get("impAssunto", []))};\n'
        f'const SAC_IMP_NIVEL={js_str(extras.get("impNivel", []))};\n'
        f'const SAC_IMP_ROWS={js_rows(imp_rows or [])};\n'
        '/* SAC_DATA_END */'
    )


def escrever_erros(caminho, erros, total_sacs):
    """Gera o arquivo de erros com os outliers excluidos do Tempo Medio."""
    from datetime import datetime
    agora = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    linhas = []
    linhas.append('=' * 78)
    linhas.append('  RELATORIO DE OUTLIERS - TEMPO MEDIO DE RESPOSTA (SAC)')
    linhas.append('=' * 78)
    linhas.append(f'Gerado em      : {agora}')
    linhas.append(f'Planilha (link): {SAC_LINK}')
    linhas.append(f'Aba            : {ABA_SAC}')
    linhas.append(f'Total de SACs  : {total_sacs}')
    linhas.append(f'Outliers       : {len(erros)} (EXCLUIDOS do calculo do Tempo Medio)')
    linhas.append('')
    linhas.append('Regra de validacao:')
    linhas.append('  - Tempo = Data de Finalizacao (col M) - Data do Atendimento (col C)')
    linhas.append('  - INVALIDO se for negativo, ou maior que '
                  f'{MAX_DELTA_DIAS} dias.')
    linhas.append('=' * 78)
    linhas.append('')
    if not erros:
        linhas.append('Nenhum outlier encontrado. Todos os tempos sao validos.')
    else:
        cab = f'{"Linha":>6} | {"Protocolo":<12} | {"Atendimento (C)":<16} | {"Finalizacao (M)":<16} | {"Dias":>8} | Motivo'
        linhas.append(cab)
        linhas.append('-' * len(cab))
        for e in sorted(erros, key=lambda x: x['linha']):
            linhas.append(
                f'{e["linha"]:>6} | {e["protocolo"]:<12} | {e["atendimento"]:<16} | '
                f'{e["finalizacao"]:<16} | {e["delta_dias"]:>8} | {e["motivo"]}'
            )
    with open(caminho, 'w', encoding='utf-8') as f:
        f.write('\n'.join(linhas) + '\n')
    print(f'  Arquivo de erros: {caminho} ({len(erros)} outliers)')


def carimbar_atualizacao(index_path):
    """Grava a data/hora da atualizacao no cabecalho do dashboard."""
    from datetime import datetime
    agora = datetime.now().strftime('%d/%m/%Y %H:%M')
    with open(index_path, 'r', encoding='utf-8') as f:
        c = f.read()
    c = re.sub(r'<!--LU-->.*?<!--/LU-->', lambda m: f'<!--LU-->{agora}<!--/LU-->', c, flags=re.DOTALL)
    with open(index_path, 'w', encoding='utf-8') as f:
        f.write(c)
    return agora


def atualizar_html(index_path, bloco):
    with open(index_path, 'r', encoding='utf-8') as f:
        conteudo = f.read()
    padrao = r'/\* SAC_DATA_START \*/.*?/\* SAC_DATA_END \*/'
    if not re.search(padrao, conteudo, re.DOTALL):
        raise ValueError('[ERRO] Marcadores SAC_DATA nao encontrados no index.html.')
    conteudo = re.sub(padrao, lambda m: bloco, conteudo, flags=re.DOTALL)  # lambda = literal
    with open(index_path, 'w', encoding='utf-8') as f:
        f.write(conteudo)


def main():
    print()
    print('=' * 50)
    print('  ATUALIZADOR - SAC')
    print('=' * 50)
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    try:
        print('\n[1/4] Baixando SharePoint (fonte 1)...')
        xlsx1 = baixar_xlsx(SAC_URL)

        print('\n[2/4] Processando fonte 1...')
        canal, reg, tipo, drows, erros, outros, extras, h1, raw1, imp_rows = processar(xlsx1)

        h2, raw2 = [], []
        print('\n[3/4] Baixando Google Sheets (fonte 2)...')
        try:
            xlsx2 = baixar_gsheets(SAC2_URL)
            mk = lambda lst: {v: i for i, v in enumerate(lst)}
            drows2, erros2, h2, raw2, imp_rows2 = processar2(
                xlsx2,
                canal,             mk(canal),
                reg,               mk(reg),
                tipo,              mk(tipo),
                extras['ent'],     mk(extras['ent']),
                extras['assunto'], mk(extras['assunto']),
                extras['prod'],    mk(extras['prod']),
                extras['uni'],     mk(extras['uni']),
                extras['impacto'], mk(extras['impacto']),
                extras['impUni'],     extras['impUniIdx'],
                extras['impAssunto'], extras['impAssuntoIdx'],
                extras['impNivel'],   extras['impNivelIdx'],
            )
            drows += drows2
            erros += erros2
            imp_rows += imp_rows2
            print(f'  Total combinado: {len(drows)} registros')
        except Exception as e2:
            print(f'  [AVISO] Fonte 2 falhou ({e2}). Continuando apenas com fonte 1.')

        escrever_erros(ERROS_TXT, erros, len(drows))

        print('\n[4/4] Atualizando index.html...')
        bloco = gerar_bloco(canal, reg, tipo, drows, outros, extras, h1, raw1, h2, raw2, imp_rows)
        atualizar_html(INDEX_HTML, bloco)
        ts = carimbar_atualizacao(INDEX_HTML)
        print(f'  Atualizado em: {ts}')

        print()
        print('=' * 50)
        print('  CONCLUIDO! index.html atualizado.')
        print('  Rode publicar.bat para enviar ao GitHub.')
        print('=' * 50)
        print()
    except Exception as e:
        print(f'\n[ERRO] {e}')
        import traceback; traceback.print_exc()


if __name__ == '__main__':
    main()
    try:
        input('Pressione ENTER para fechar...')
    except EOFError:
        pass
