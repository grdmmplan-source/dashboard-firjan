# -*- coding: utf-8 -*-
"""
Atualizador Dashboard Firjan - Campanha Colonia Inverno 2026
Le a base de clientes (Google Sheets) + a base de ligacoes (Discagem) e
atualiza o bloco 'colonia_inverno' no index.html.

Base de clientes -> card "Cliente na Base" e grafico de Distribuicao,
aba "Por Unidade" (coluna K - Unidade de Interesse).
Base de ligacoes (Discagem) -> Tentativas, Interessados, Decisor,
Conversao, Media, Distribuicao "Por Status" e Evolucao Diaria.
"""

import csv
import glob
import io
import os
import re
import urllib.request
from collections import Counter, defaultdict
from datetime import datetime, timedelta

# ═══════════════════════════════════════════════════════════
# CONFIGURAÇÃO
# ═══════════════════════════════════════════════════════════

SHEET_ID   = '1nKJDb-X5YTUEh4IMop30mgQerGk5FeMP'
GID        = '116593147'
INDEX_HTML = r'index.html'
DEPARA_PATH = r'Arquivos\bases_apoio\tab_de-para.xlsx'
PASTA_DISCAGEM = r'Arquivos\nao_atualizaveis\Ativo_Colônia de Férias'

COL_UNIDADE = 10  # K - Unidade de Interesse (base de clientes)
COL_MOTIVO  = 3   # D - Motivo do não interesse (base de clientes)

STATUS_MAP = {}
LABELS_NAO_DECISOR = {'Telefonia', 'Tentativa', 'Engano', 'Alo'}
LABEL_INTERESSADO  = 'Interessado'

MES_PT = {1:'Jan',2:'Fev',3:'Mar',4:'Abr',5:'Mai',6:'Jun',
          7:'Jul',8:'Ago',9:'Set',10:'Out',11:'Nov',12:'Dez'}

def norm_txt(s):
    """Maiuscula, sem acento - para comparar tabulacoes brutas."""
    import unicodedata
    b = unicodedata.normalize('NFKD', str(s).strip().upper())
    return ''.join(c for c in b if not unicodedata.combining(c))

# Grafico "Contatos de Sucesso" (visao Por Status): tabulacoes brutas (antes do de-para)
# consideradas sucesso. Comparadas via norm_txt.
SUCESSO_LABELS = ['JÁ INSCRITO', 'MATRÍCULA ONLINE', 'Não tem interesse', 'RETORNAR', 'INTERESSADO']
SUCESSO_MAP = {norm_txt(s): s for s in SUCESSO_LABELS}

# Grafico "Tentativas de Contato Sem Sucesso": tabulacoes sem interacao com o operador
# sao agrupadas sob o rotulo unico "Tentativa"
SEM_OPERADOR_LABELS = ['Atendido', 'FORA DE ÁREA / CX MENSAGENS', 'Ligação Muda',
                        'Não Atendeu', 'Ocupado', 'TEL NÃO ATENDE / OCUPADO']
SEM_OPERADOR_LABEL_CANON = 'Tentativa'
SEM_OPERADOR_NORM = {norm_txt(s) for s in SEM_OPERADOR_LABELS}


# ═══════════════════════════════════════════════════════════
# FUNÇÕES AUXILIARES
# ═══════════════════════════════════════════════════════════

def fmt_num(n):
    return f"{n:,}".replace(",", ".")


def fmt_pct(n):
    return f"{n:.2f}%".replace(".", ",")


def fmt_dec(n):
    return f"{n:.2f}".replace(".", ",")


def norm_tel(t):
    return re.sub(r'\D', '', str(t)) if t else ''


def to_datetime(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, (int, float)) and val > 0:
        try:
            return datetime(1899, 12, 30) + timedelta(days=float(val))
        except Exception:
            return None
    return None


def encontrar_unico_xlsx(pasta):
    arquivos = [a for a in glob.glob(os.path.join(pasta, '*.xlsx')) if not os.path.basename(a).startswith('~$')]
    if not arquivos:
        raise FileNotFoundError(f'[ERRO] Nenhum .xlsx encontrado em {pasta}')
    return sorted(arquivos)[-1]


def ler_depara(caminho):
    mapa = {}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
        ws = wb['status'] if 'status' in wb.sheetnames else wb.active
        first = True
        for row in ws.iter_rows(values_only=True):
            if first:
                first = False
                continue
            if row[0] and row[1]:
                mapa[str(row[0]).strip()] = str(row[1]).strip()
        wb.close()
        print(f'  De-para: {len(mapa)} entradas')
    except FileNotFoundError:
        print(f'  [AVISO] De-para nao encontrado: {caminho}')
    return mapa


def normalizar_status(raw):
    if raw is None:
        return None
    s = str(raw).strip()
    for k, v in STATUS_MAP.items():
        if s.upper() == k.upper():
            return v
    return s


# ═══════════════════════════════════════════════════════════
# BASE DE CLIENTES (Google Sheets)
# ═══════════════════════════════════════════════════════════

def baixar_csv(sheet_id, gid=None):
    url = f'https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv'
    if gid:
        url += f'&gid={gid}'
    resp = urllib.request.urlopen(url, timeout=60)
    conteudo = resp.read().decode('utf-8-sig')
    return list(csv.reader(io.StringIO(conteudo)))


def calcular_base():
    print('  Baixando base de clientes Colônia Inverno 2026 (Google Sheets)...')
    linhas = baixar_csv(SHEET_ID, GID)
    dados = linhas[1:]

    clientes = 0
    uni_counter = Counter()
    motivo_counter = Counter()
    for r in dados:
        if not any(c.strip() for c in r):
            continue
        clientes += 1
        uni = r[COL_UNIDADE].strip() if len(r) > COL_UNIDADE else ''
        if uni:
            uni_counter[uni] += 1
        motivo = r[COL_MOTIVO].strip() if len(r) > COL_MOTIVO else ''
        if motivo:
            motivo_counter[motivo] += 1

    uni_items = uni_counter.most_common()
    motivo_items = motivo_counter.most_common()
    print(f'  Clientes na base: {clientes}')
    print(f'  Com unidade de interesse: {sum(uni_counter.values())} | Unidades distintas: {len(uni_items)}')
    print(f'  Com motivo de nao interesse: {sum(motivo_counter.values())} | Motivos distintos: {len(motivo_items)}')

    return {
        'clientes':     clientes,
        'uniLabels':    [u[0] for u in uni_items],
        'uniData':      [u[1] for u in uni_items],
        'motivoLabels': [m[0] for m in motivo_items],
        'motivoData':   [m[1] for m in motivo_items],
    }


# ═══════════════════════════════════════════════════════════
# BASE DE LIGAÇÕES (Discagem)
# ═══════════════════════════════════════════════════════════

def calcular_discagem(caminho):
    """Le o Discagem_Ativo_Colônia de Férias.xlsx.
    Cada linha com DATA (col A) = 1 tentativa de ligacao.
    Status: col L (STATUS_NEGOCIO); se vazia, usa col K (STATUS).
    """
    import openpyxl
    print(f'  Lendo Discagem: {os.path.basename(caminho)}')
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    aba = next((s for s in wb.sheetnames if s.startswith('chamadas_')), wb.sheetnames[0])
    ws = wb[aba]

    data_idx = 0    # A DATA
    orig_idx = 7    # H ORIGEM
    dest_idx = 8    # I DESTINO
    tipo_idx = 9    # J TIPO
    st_idx   = 10   # K STATUS
    sn_idx   = 11   # L STATUS_NEGOCIO

    total_tent = 0
    status_counter = Counter()
    raw_por_label  = defaultdict(set)
    naosucesso_counter = Counter()
    raw_por_label_ns   = defaultdict(set)
    decisor_count = 0
    tel_interesse = set()
    data_min = data_max = None

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        dt = to_datetime(row[data_idx])
        if not dt:
            continue
        total_tent += 1
        if data_min is None or dt < data_min:
            data_min = dt
        if data_max is None or dt > data_max:
            data_max = dt

        tipo = str(row[tipo_idx]).strip().upper() if len(row) > tipo_idx and row[tipo_idx] else ''
        if tipo == 'DISCADOR':
            tel = norm_tel(row[orig_idx])
        elif tipo == 'SAINTE':
            tel = norm_tel(row[dest_idx])
        else:
            tel = norm_tel(row[orig_idx]) or norm_tel(row[dest_idx])

        sn_raw = row[sn_idx] if len(row) > sn_idx else None
        st_raw = row[st_idx] if len(row) > st_idx else None
        raw    = str(sn_raw).strip() if sn_raw else (str(st_raw).strip() if st_raw else '')
        raw_n  = norm_txt(raw) if raw else ''
        label  = normalizar_status(raw) if raw else None

        if raw_n in SUCESSO_MAP:
            canon = SUCESSO_MAP[raw_n]
            status_counter[canon] += 1
            raw_por_label[canon].add(raw)
            decisor_count += 1
        elif raw:
            ns_label = SEM_OPERADOR_LABEL_CANON if raw_n in SEM_OPERADOR_NORM else raw
            naosucesso_counter[ns_label] += 1
            raw_por_label_ns[ns_label].add(raw)
        else:
            naosucesso_counter['Tentativas de Contato Sem Sucesso'] += 1
            raw_por_label_ns['Tentativas de Contato Sem Sucesso'].add('(vazio)')

        # Interessado = label == 'Interessado' (nao alterado por esta mudanca)
        if label == LABEL_INTERESSADO:
            tel_interesse.add(tel)

    wb.close()

    st_items = status_counter.most_common()
    st_tooltips = [', '.join(sorted(raw_por_label.get(s[0], set()))) for s in st_items]
    ns_items = naosucesso_counter.most_common()
    ns_tooltips = [', '.join(sorted(raw_por_label_ns.get(s[0], set()))) for s in ns_items]

    print(f'  Total Tentativas: {total_tent}')
    print(f'  Decisor: {decisor_count} | Interessados: {len(tel_interesse)}')
    print(f'  Status: {dict(st_items[:5])} ...')

    return {
        'tentativas':    total_tent,
        'decisor':       decisor_count,
        'interessados':  len(tel_interesse),
        'naosucessoLabels':   [s[0] for s in ns_items],
        'naosucessoData':     [s[1] for s in ns_items],
        'naosucessoTooltips': ns_tooltips,
        'dataMin': data_min, 'dataMax': data_max,
        'statusLabels':  [s[0] for s in st_items],
        'statusData':    [s[1] for s in st_items],
        'statusTooltips': st_tooltips,
    }


# ═══════════════════════════════════════════════════════════
# GERAÇÃO DO BLOCO JS
# ═══════════════════════════════════════════════════════════

def js_str(lst):
    return '[' + ','.join(f"'{str(v).replace(chr(39), chr(92)+chr(39))}'" for v in lst) + ']'


def js_num(lst):
    return '[' + ','.join(str(v) for v in lst) + ']'


def gerar_bloco(base, disc):
    clientes = base['clientes']
    tent     = disc['tentativas']
    decisor  = disc['decisor']
    interess = disc['interessados']
    taxa     = (interess / decisor * 100) if decisor > 0 else 0
    media    = (tent / clientes) if clientes > 0 else 0

    if disc.get('dataMin') and disc.get('dataMax'):
        dmin, dmax = disc['dataMin'], disc['dataMax']
        periodo = f"{dmin.day:02d}/{MES_PT[dmin.month]} — {dmax.day:02d}/{MES_PT[dmax.month]}"
    else:
        periodo = ''
    status_labels = disc['statusLabels'] or ['Sem dados']
    status_data   = disc['statusData'] or [0]
    ns_labels     = disc['naosucessoLabels'] or ['Sem dados']
    ns_data       = disc['naosucessoData'] or [0]

    return f"""  /* COLONIA_INVERNO_START */
  colonia_inverno: {{
    label: '— Colônia Inverno 2026', desc: 'Campanha Colônia Inverno 2026 — dados filtrados', periodo: '{periodo}',
    empresas: '{fmt_num(clientes)}', empresasLabel: '🧒 Cliente na Base',
    mediaLabel: '🔁 Média Tentativas/Cliente', mediaSub: 'por cliente',
    tentativas: '{fmt_num(tent)}', interessados: '{fmt_num(interess)}', conversao: '{fmt_pct(taxa)}',
    decisor: '{fmt_num(decisor)}', decisorLabel: '👤 Contatos de Sucesso', decisorSub: 'Apenas Colônia Inverno 2026', media: '{fmt_dec(media)}', trend: '',
    distTitle: 'Contatos de Sucesso',
    statusLabels: {js_str(status_labels)}, statusData: {js_num(status_data)}, statusColors:null,
    statusTooltips: {js_str(disc['statusTooltips'])},
    evoTitle: 'Tentativas de Contato Sem Sucesso', evoBar: true,
    naosucessoLabels: {js_str(ns_labels)}, naosucessoData: {js_num(ns_data)},
    naosucessoTooltips: {js_str(disc['naosucessoTooltips'])},
    evolucaoLabels: [], tentDia: [], convDia: [],
    showWpp: false,
    wppTitle: '', wppDesc: '',
    wppKpiLabels: [], wppListLabels: [], wppPieLabels: [],
    wppEnv:'-', wppResp:'-', wppTaxa:'-', wppSem:'-', wppInfo:'-', wppEmail:'-', wppPie:[0,1],
    distToggle: true,
    uniLabels: {js_str(base['uniLabels'])},
    uniData: {js_num(base['uniData'])},
    showMotivo: true,
    motivoLabels: {js_str(base['motivoLabels'])},
    motivoData: {js_num(base['motivoData'])}
  }},
  /* COLONIA_INVERNO_END */"""


def atualizar_html(index_path, blocos):
    with open(index_path, 'r', encoding='utf-8') as f:
        conteudo = f.read()
    for marcador, novo_bloco in blocos.items():
        padrao = f'/\\* {marcador}_START \\*/.*?/\\* {marcador}_END \\*/'
        if not re.search(padrao, conteudo, re.DOTALL):
            raise ValueError(f'[ERRO] Marcadores {marcador} nao encontrados no index.html.')
        conteudo = re.sub(padrao, novo_bloco, conteudo, flags=re.DOTALL)
    with open(index_path, 'w', encoding='utf-8') as f:
        f.write(conteudo)


# ═══════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════

def main():
    print()
    print('=' * 50)
    print('  ATUALIZADOR — Colônia Inverno 2026')
    print('=' * 50)
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    try:
        STATUS_MAP.update(ler_depara(DEPARA_PATH))
        base = calcular_base()
        disc = calcular_discagem(encontrar_unico_xlsx(PASTA_DISCAGEM))
        bloco = {'COLONIA_INVERNO': gerar_bloco(base, disc)}
        atualizar_html(INDEX_HTML, bloco)
        print()
        print('=' * 50)
        print('  CONCLUIDO! index.html atualizado.')
        print('  Rode publicar.bat para enviar ao GitHub.')
        print('=' * 50)
        print()
    except Exception as e:
        print(f'\n[ERRO] {e}')
        import traceback; traceback.print_exc()


if __name__ == '__main__':
    main()
