# -*- coding: utf-8 -*-
"""
Atualizador Dashboard Firjan - Modulo Receptivo
Le o arquivo Retorno_RECEPTIVO.xlsx (aba BSales2) e injeta os dados brutos no
index.html. Os filtros (Data, Canal, Entidade, Unidade) e os KPIs sao
recalculados no navegador.

Aba: "BSales2"
Colunas usadas (0-based):
  D(3)  = Data/Hora de abertura -> filtro Data (calendario)
  E(4)  = Origem do caso        -> filtro Canal + grafico "Distribuicao por Canal"
  G(6)  = Entidade              -> filtro (oculta caixa postal, ligacao muda,
                                   queda de ligacao, spam e vazio no dropdown)
  O(14) = Unidade               -> filtro

Total de Atendimentos = quantidade de linhas do BSales2.
"""

import openpyxl
import re
import glob
import os
import io
import urllib.request
import http.cookiejar
from datetime import datetime, timedelta
import atualizar_csat

# ═══════════════════════════════════════════════════════════
# CONFIGURAÇÃO
# ═══════════════════════════════════════════════════════════

PASTA       = r'Arquivos\atualizaveis'
PREFIXO     = 'Retorno_RECEPTIVO'
ABA         = 'BSales2'
INDEX_HTML  = r'index.html'

# --- Monitoria (IEC) ---
MONITORIA_PREFIXO = 'Retorno_monitoria_analitico'
MON_CAMPANHA = 2    # C  CAMPANHA (Telefone vs demais=Digital)
MON_NOTA     = 13   # N  NOTA  (IEC = notas zeradas / total)

COL_DATA  = 3    # D  Data/Hora de abertura
COL_CANAL = 4    # E  Origem do caso
COL_ENT   = 23   # X  Entidade2 -> filtro Entidade
COL_CLASS = 11   # L  Classificacao do Ticket -> IAR
COL_UNI   = 22   # W  Unidade2 -> filtro Unidade
COL_REG2  = 24   # Y  Regional2 -> filtro Regional
COL_SEG   = 15   # P  Segmento
COL_PES   = 16   # Q  Pesquisa

# --- Aba "Redes Sociais" (no arquivo SharePoint do SAC) ---
REDES_URL = ('https://ddmadvbr-my.sharepoint.com/:x:/g/personal/'
             'fernanda_castro_grupoddm_com_br/'
             'IQCobMqlyAWfT7aBxT5EGURCAaAxahCZQjDD0eCyvVChMv0?download=1')
RS_ASSUNTO  = 4   # E  Assunto
RS_UNIDADE  = 5   # F  Unidade
RS_ENT      = 6   # G  Entidade -> filtro Entidade
RS_REGIONAL = 7   # H  Regional -> filtro Regional
RS_DATA     = 9   # J  Resolvido (data)
RS_CANAL    = 10  # K  Observacao -> Canal (Instagram/Facebook/Messenger)

# --- Arquivo "Autonomia e Renda" (SharePoint, separado por abas/meses) ---
AUTONOMIA_URL = ('https://ddmadvbr-my.sharepoint.com/:x:/g/personal/'
                 'fernanda_castro_grupoddm_com_br/'
                 'IQCwxtAUa9lQS4BgM7YdpFrxAQd1Jys7p98gbp4QY43dnSg?download=1')
AUTONOMIA_CANAL_LABEL = 'Autonomia e Renda'
AUT_HDR_DATA = 'DATA INICIAL'   # coluna de data (achada pelo nome)
AUT_HDR_CAT  = 'CATEGORIA'      # coluna do grafico Servicos Mais Procurados


# --- Aba BASE DISCADOR 1 (acionamentos do discador) ---
DISC_ABA      = 'BASE DISCADOR 1'
DISC_FILAS    = ('3000', '4000')   # filas consideradas (coluna A)
DC_FILA  = 0     # A  FILA
DC_DATA  = 1     # B  DATA  (formato "DD/MM HH:MM")
DC_ATEND = 3     # D  ATEND.
DC_ABAND = 5     # F  ABAND
DC_NSREC = 9     # J  N.S. RECEB. (ligacoes atendidas ate 30s) -> ICT
DC_TMA   = 14   # O  TMADURACAO -> TMA Voz = soma(col O) / soma(ATEND)
DC_TME   = 11   # L  TMEDURACAO -> TME Voz = soma(col L) / soma(ATEND)


# ═══════════════════════════════════════════════════════════
# FUNÇÕES
# ═══════════════════════════════════════════════════════════

def encontrar_arquivo(pasta, prefixo):
    padrao = os.path.join(pasta, f'{prefixo}*.xlsx')
    arquivos = [a for a in glob.glob(padrao) if not os.path.basename(a).startswith('~$')]
    if not arquivos:
        raise FileNotFoundError(f'[ERRO] Arquivo nao encontrado: {prefixo}*.xlsx em {pasta}')
    return sorted(arquivos)[-1]


def to_dt(val):
    if isinstance(val, datetime):
        return val
    if isinstance(val, (int, float)) and val > 0:
        try:
            return datetime(1899, 12, 30) + timedelta(days=float(val))
        except Exception:
            return None
    return None


def cel(row, idx):
    if len(row) <= idx:
        return None
    v = row[idx]
    return v


def txt(v):
    return str(v).strip() if v is not None else ''


def num(v):
    try:
        return float(str(v).replace(',', '.'))
    except Exception:
        return 0.0


def norm_txt(s):
    """Chave de agrupamento: minuscula, sem acento."""
    import unicodedata
    b = unicodedata.normalize('NFKD', str(s).strip().lower())
    return ''.join(c for c in b if not unicodedata.combining(c))


def parse_any_date(v):
    """datetime -> yyyymmdd; texto 'dd/mm/yyyy' -> yyyymmdd; senao 0."""
    d = to_dt(v)
    if d:
        return d.year * 10000 + d.month * 100 + d.day
    m = re.match(r'(\d{1,2})/(\d{1,2})/(\d{2,4})', str(v).strip())
    if m:
        dd, mo, yy = m.groups()
        yy = int(yy)
        if yy < 100:
            yy += 2000
        return yy * 10000 + int(mo) * 100 + int(dd)
    return 0


def ler_iec(pasta):
    """IEC = notas zeradas / total, separado em Telefone e Digital (resto).
    Retorna (iec_tel, iec_dig) em % (0 se sem dados)."""
    tel = [0, 0]   # [zeradas, total]
    dig = [0, 0]
    try:
        cam = encontrar_arquivo(pasta, MONITORIA_PREFIXO)
        wb = openpyxl.load_workbook(cam, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        for i, r in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            if not any(c is not None and str(c).strip() for c in r):
                continue
            camp = txt(cel(r, MON_CAMPANHA)).lower()
            alvo = tel if camp == 'telefone receptivo' else dig
            alvo[1] += 1
            nota = cel(r, MON_NOTA)
            try:
                if nota is not None and float(str(nota).replace(',', '.')) == 0:
                    alvo[0] += 1
            except ValueError:
                pass
        wb.close()
        iec_tel = (tel[0] / tel[1] * 100) if tel[1] else 0
        iec_dig = (dig[0] / dig[1] * 100) if dig[1] else 0
        print(f'  IEC Telefone: {tel[0]}/{tel[1]} = {iec_tel:.1f}% | Digital: {dig[0]}/{dig[1]} = {iec_dig:.1f}%')
        return iec_tel, iec_dig
    except Exception as e:
        print(f'  [AVISO] IEC (monitoria) nao carregado: {e}')
        return 0, 0


def baixar_sharepoint(url):
    """Baixa o .xlsx do SharePoint (link anonimo) via cookie jar."""
    cj = http.cookiejar.CookieJar()
    op = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(cj))
    op.addheaders = [('User-Agent', 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)')]
    resp = op.open(url, timeout=60)
    return resp.read()


def to_seconds(v):
    """Converte celula de duracao em segundos; None se nao houver tempo."""
    from datetime import time as _t
    if isinstance(v, timedelta):
        return v.total_seconds()
    if isinstance(v, _t):
        return v.hour * 3600 + v.minute * 60 + v.second
    if isinstance(v, (int, float)) and v > 0:
        # Se o valor for grande (>= 1), provavelmente ja esta em segundos
        if v >= 1:
            return float(v)
        return float(v) * 86400  # fracao de dia do Excel
    if isinstance(v, str):
        parts = v.strip().split(':')
        try:
            if len(parts) == 3:
                return int(parts[0]) * 3600 + int(parts[1]) * 60 + float(parts[2])
            if len(parts) == 2:
                return int(parts[0]) * 60 + float(parts[1])
        except (ValueError, IndexError):
            pass
    return None


def parse_disc_data(v, ref_year):
    """Converte 'DD/MM HH:MM' -> int yyyymmdd usando o ano de referencia."""
    m = re.match(r'(\d{1,2})/(\d{1,2})', str(v).strip())
    if not m:
        return 0
    d, mo = m.groups()
    return ref_year * 10000 + int(mo) * 100 + int(d)


def processar(caminho):
    print(f'  Lendo: {os.path.basename(caminho)} | aba: {ABA}')
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    if ABA not in wb.sheetnames:
        raise ValueError(f'Aba "{ABA}" nao encontrada. Abas: {wb.sheetnames}')
    ws = wb[ABA]

    canal_list, ent_list, uni_list, assunto_list, reg_list, seg_list, pes_list = [], [], [], [], [], [], []
    canal_idx, ent_idx, uni_idx, assunto_idx, reg_idx, seg_idx, pes_idx = {}, {}, {}, {}, {}, {}, {}

    def get_idx(val, lst, mp):
        v = txt(val)
        if v not in mp:
            mp[v] = len(lst)
            lst.append(v)
        return mp[v]

    data_rows = []
    total = 0
    max_dt = 0
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # cabecalho
        if not any(c is not None and str(c).strip() for c in r):
            continue  # linha totalmente vazia
        total += 1
        c = to_dt(cel(r, COL_DATA))
        dt = (c.year * 10000 + c.month * 100 + c.day) if c else 0
        if dt > max_dt:
            max_dt = dt
        ci = get_idx(cel(r, COL_CANAL), canal_list, canal_idx)
        gi = get_idx(cel(r, COL_ENT),   ent_list,   ent_idx)
        oi = get_idx(cel(r, COL_UNI),   uni_list,   uni_idx)
        ri = get_idx(cel(r, COL_REG2),  reg_list,   reg_idx)   # Regional (W = Regional2)
        classif = txt(cel(r, COL_CLASS))             # Classificacao do Ticket
        hc = 1 if classif else 0                      # tem classificacao? (IAR)
        segi = get_idx(cel(r, COL_SEG), seg_list, seg_idx)
        pesi = get_idx(cel(r, COL_PES), pes_list, pes_idx)
        # [dt, canal, entidade, unidade, hc, assunto, regional, src(0=BSales2), segmento, pesquisa]
        data_rows.append([dt, ci, gi, oi, hc, classif or None, ri, 0, segi, pesi])

    # Ano de referencia (a aba do discador nao tem ano na data)
    ref_year = (max_dt // 10000) if max_dt else 2026

    # --- BASE DISCADOR 1 (fila 3000): agrega ATEND e ABAND por dia ---
    disc_daily = {}
    if DISC_ABA in wb.sheetnames:
        wsd = wb[DISC_ABA]
        for i, r in enumerate(wsd.iter_rows(values_only=True)):
            if i == 0:
                continue
            if txt(cel(r, DC_FILA)) not in DISC_FILAS:
                continue
            dt = parse_disc_data(cel(r, DC_DATA), ref_year)
            if not dt:
                continue
            atend = num(cel(r, DC_ATEND))
            aband = num(cel(r, DC_ABAND))
            nsrec = num(cel(r, DC_NSREC))
            tma_d = to_seconds(cel(r, DC_TMA)) or 0.0
            tme_d = to_seconds(cel(r, DC_TME)) or 0.0
            if dt not in disc_daily:
                disc_daily[dt] = [0.0, 0.0, 0.0, 0.0, 0.0]
            disc_daily[dt][0] += atend
            disc_daily[dt][1] += aband
            disc_daily[dt][2] += nsrec
            disc_daily[dt][3] += tma_d
            disc_daily[dt][4] += tme_d
    else:
        print(f'  [AVISO] Aba "{DISC_ABA}" nao encontrada.')

    disc2_daily = {}  # DISCADOR 2 desconsiderado

    # --- CSAT (satisfacao): consolidado de 6 Google Sheets ---
    print('  Carregando CSAT de Google Sheets...')
    csat_rows = atualizar_csat.main()
    extra_canais = []
    for row in csat_rows:
        canal = row[1]  # [dt, canal, bons, tot, bons_ces]
        if canal and canal not in canal_list and canal not in extra_canais:
            extra_canais.append(canal)

    # --- Redes Sociais (SharePoint do SAC): canal = Observacao (K); regional = Entidade (G) ---
    n_redes = 0
    try:
        rb = baixar_sharepoint(REDES_URL)
        wbr = openpyxl.load_workbook(io.BytesIO(rb), read_only=True, data_only=True)
        aba_r = next((s for s in wbr.sheetnames if 'edes' in s.lower()), None)
        if aba_r:
            wsr = wbr[aba_r]
            for i, r in enumerate(wsr.iter_rows(values_only=True)):
                if i == 0:
                    continue
                if not any(c is not None and str(c).strip() for c in r):
                    continue
                c = to_dt(cel(r, RS_DATA))
                dt = (c.year * 10000 + c.month * 100 + c.day) if c else 0
                ci_rs = get_idx(cel(r, RS_CANAL), canal_list, canal_idx)
                gi_rs = get_idx(cel(r, RS_ENT),      ent_list,   ent_idx)
                ri_rs = get_idx(cel(r, RS_REGIONAL), reg_list,   reg_idx)
                oi    = get_idx(cel(r, RS_UNIDADE),  uni_list,   uni_idx)
                araw  = txt(cel(r, RS_ASSUNTO))
                data_rows.append([dt, ci_rs, gi_rs, oi, 0, araw or None, ri_rs, 1, -1, -1])
                n_redes += 1
        else:
            print('  [AVISO] Aba "Redes Sociais" nao encontrada no SharePoint.')
        wbr.close()
    except Exception as e:
        print(f'  [AVISO] Redes Sociais nao carregada: {e}')

    # --- Autonomia e Renda (SharePoint, varias abas/meses) -> canal "Autonomia e Renda" ---
    n_aut = 0
    try:
        ab = baixar_sharepoint(AUTONOMIA_URL)
        wba = openpyxl.load_workbook(io.BytesIO(ab), read_only=True, data_only=True)
        ci_aut = get_idx(AUTONOMIA_CANAL_LABEL, canal_list, canal_idx)
        gi_vazio = get_idx('', ent_list, ent_idx)
        oi_vazio = get_idx('', uni_list, uni_idx)
        ri_vazio = get_idx('', reg_list, reg_idx)   # Autonomia: sem regional
        for aba in wba.sheetnames:
            ws = wba[aba]
            hdr_i = None
            colmap = {}
            for i, r in enumerate(ws.iter_rows(values_only=True, max_row=10)):
                ups = [txt(c).upper() for c in r]
                if AUT_HDR_DATA in ups:
                    hdr_i = i
                    colmap = {txt(c).upper(): j for j, c in enumerate(r) if txt(c)}
                    break
            if hdr_i is None:
                continue  # aba sem dados
            di = colmap.get(AUT_HDR_DATA)
            cat = colmap.get(AUT_HDR_CAT)
            for i, r in enumerate(ws.iter_rows(values_only=True)):
                if i <= hdr_i:
                    continue
                dval = r[di] if (di is not None and len(r) > di) else None
                if dval is None or not str(dval).strip():
                    continue  # so conta linha com Data Inicial
                dt = parse_any_date(dval)
                catv = txt(r[cat]) if (cat is not None and len(r) > cat) else ''
                data_rows.append([dt, ci_aut, gi_vazio, oi_vazio, 0, catv or None, ri_vazio, 2, -1, -1])
                n_aut += 1
        wba.close()
    except Exception as e:
        print(f'  [AVISO] Autonomia e Renda nao carregada: {e}')

    # --- Padroniza assuntos/categorias (case+acento) -> forma mais frequente ---
    from collections import Counter as _C
    raw_cnt = _C(r[5] for r in data_rows if isinstance(r[5], str) and r[5])
    key_forms = {}
    for raw, c in raw_cnt.items():
        key_forms.setdefault(norm_txt(raw), _C())[raw] += c
    canon_ass = {k: cc.most_common(1)[0][0] for k, cc in key_forms.items()}
    assunto_list, assunto_idx = [], {}
    for r in data_rows:
        a = r[5]
        if isinstance(a, str) and a:
            disp = canon_ass[norm_txt(a)]
            if disp not in assunto_idx:
                assunto_idx[disp] = len(assunto_list)
                assunto_list.append(disp)
            r[5] = assunto_idx[disp]
        else:
            r[5] = -1

    disc_rows = [[dt, int(v[0]), int(v[1]), int(v[2]), round(v[3]), round(v[4])] for dt, v in sorted(disc_daily.items())]
    tot_atend = sum(v[1] for v in disc_rows)
    tot_aband = sum(v[2] for v in disc_rows)
    tot_nsrec = sum(v[3] for v in disc_rows)
    tel_bsales = sum(1 for r in data_rows if canal_list[r[1]] == 'Telefone')
    ial = (tot_aband / tot_atend * 100) if tot_atend else 0
    ict = (tot_nsrec / tot_atend * 100) if tot_atend else 0
    s_tma = sum(v[4] for v in disc_rows)
    s_tme = sum(v[5] for v in disc_rows)
    tma = (s_tma / tot_atend) if tot_atend else 0
    tme = (s_tme / tot_atend) if tot_atend else 0
    cs_bons = sum(r[2] for r in csat_rows); cs_tot = sum(r[3] for r in csat_rows)
    bsales_rows = [r for r in data_rows if r[7] == 0]   # src==0 -> BSales2
    iar = (sum(r[4] for r in bsales_rows) / len(bsales_rows) * 100) if bsales_rows else 0
    csat = (cs_bons / cs_tot * 100) if cs_tot else 0

    wb.close()

    iec_tel, iec_dig = ler_iec(PASTA)

    print(f'  Total Atendimentos (BSales2): {total}')
    print(f'  Canais: {len(canal_list)} | Entidades: {len(ent_list)} | Unidades: {len(uni_list)}')
    print(f'  Discador filas {DISC_FILAS}: dias={len(disc_rows)} | ATEND={tot_atend} | ABAND={tot_aband}')
    print(f'  Telefone na BSales2: {tel_bsales} | Diferenca somada: {max(0, tot_atend - tel_bsales)}')
    print(f'  IAL (ABAND/ATEND): {ial:.2f}% | ICT (N.S.RECEB/ATEND): {ict:.2f}%')
    print(f'  TMA Voz: {tma:.0f}s | TME Voz: {tme:.0f}s (ATEND={tot_atend})')
    print(f'  IAR (com Classificacao/total BSales2): {iar:.2f}%')
    print(f'  CSAT: bons={cs_bons} total={cs_tot} -> {csat:.2f}% | canais extras: {extra_canais}')
    print(f'  Redes Sociais: {n_redes} atendimentos')
    print(f'  Autonomia e Renda: {n_aut} atendimentos')
    print(f'  Assuntos/Categorias distintos: {len(assunto_list)}')
    print(f'  Total geral (BSales2 + Redes + Autonomia): {len(data_rows)}')
    print(f'  Regionais: {len([x for x in reg_list if x])}')
    return (canal_list, ent_list, uni_list, data_rows, disc_rows,
            csat_rows, extra_canais, assunto_list, reg_list, iec_tel, iec_dig, seg_list, pes_list)


def gerar_bloco(canal_list, ent_list, uni_list, data_rows, disc_rows=None,
                csat_rows=None, extra_canais=None, assunto_list=None, reg_list=None,
                iec_tel=0, iec_dig=0, seg_list=None, pes_list=None):
    def js_str(lst):
        return '[' + ','.join("'" + str(v).replace('\\', '\\\\').replace("'", "\\'") + "'" for v in lst) + ']'
    def js_num_rows(rows):
        return '[' + ','.join('[' + ','.join(str(c) for c in r) + ']' for r in rows) + ']'
    def js_csat_rows(rows):
        # [dt, 'canal', bons(D+F+H), total(D+F+H), bons_H]
        out = []
        for r in (rows or []):
            canal = "'" + str(r[1]).replace('\\', '\\\\').replace("'", "\\'") + "'"
            out.append(f'[{r[0]},{canal},{r[2]},{r[3]},{r[4]}]')
        return '[' + ','.join(out) + ']'

    return (
        '/* RECEP_DATA_START */\n'
        f'const RECEP_CANAL={js_str(canal_list)};\n'
        f'const RECEP_CANAL_EXTRA={js_str(extra_canais or [])};\n'
        f'const RECEP_ENT={js_str(ent_list)};\n'
        f'const RECEP_REG={js_str(reg_list or [])};\n'
        f'const RECEP_UNI={js_str(uni_list)};\n'
        f'const RECEP_ASSUNTO={js_str(assunto_list or [])};\n'
        f'const RECEP_SEG={js_str(seg_list or [])};\n'
        f'const RECEP_PES={js_str(pes_list or [])};\n'
        f'const RECEP_ROWS={js_num_rows(data_rows)};\n'
        f'const DISC_ROWS={js_num_rows(disc_rows or [])};\n'
        f'const CSAT_ROWS={js_csat_rows(csat_rows)};\n'
        f'const IEC_TEL=null;\n'
        f'const IEC_DIG=null;\n'
        '/* RECEP_DATA_END */'
    )


def atualizar_html(index_path, bloco):
    with open(index_path, 'r', encoding='utf-8') as f:
        conteudo = f.read()
    padrao = r'/\* RECEP_DATA_START \*/.*?/\* RECEP_DATA_END \*/'
    if not re.search(padrao, conteudo, re.DOTALL):
        raise ValueError('[ERRO] Marcadores RECEP_DATA nao encontrados no index.html.')
    conteudo = re.sub(padrao, lambda m: bloco, conteudo, flags=re.DOTALL)  # lambda = literal
    with open(index_path, 'w', encoding='utf-8') as f:
        f.write(conteudo)


def carimbar_atualizacao(index_path):
    agora = datetime.now().strftime('%d/%m/%Y %H:%M')
    with open(index_path, 'r', encoding='utf-8') as f:
        c = f.read()
    c = re.sub(r'<!--LU-->.*?<!--/LU-->', lambda m: f'<!--LU-->{agora}<!--/LU-->', c, flags=re.DOTALL)
    with open(index_path, 'w', encoding='utf-8') as f:
        f.write(c)
    return agora


def main():
    print()
    print('=' * 50)
    print('  ATUALIZADOR - Receptivo')
    print('=' * 50)
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    try:
        print('\n[1/2] Processando dados...')
        caminho = encontrar_arquivo(PASTA, PREFIXO)
        canal, ent, uni, drows, disc, csat, extra, assunto, reg, iecT, iecD, seg, pes = processar(caminho)  # 13 valores

        print('\n[2/2] Atualizando index.html...')
        bloco = gerar_bloco(canal, ent, uni, drows, disc, csat, extra, assunto, reg, iecT, iecD, seg, pes)
        atualizar_html(INDEX_HTML, bloco)
        ts = carimbar_atualizacao(INDEX_HTML)
        print(f'  Atualizado em: {ts}')

        print()
        print('=' * 50)
        print('  CONCLUIDO! index.html atualizado.')
        print('  Rode publicar.bat para enviar ao GitHub.')
        print('=' * 50)
        print()
    except Exception as e:
        print(f'\n[ERRO] {e}')
        import traceback; traceback.print_exc()


if __name__ == '__main__':
    main()
    try:
        input('Pressione ENTER para fechar...')
    except EOFError:
        pass
