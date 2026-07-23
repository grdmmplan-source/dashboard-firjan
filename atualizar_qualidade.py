# -*- coding: utf-8 -*-
"""
Atualizador Dashboard Firjan - Secao Qualidade (Receptivo)
Le o Relatorio Monitoria_Firjan.xlsx e atualiza o bloco QUALIDADE no index.html.

Aba "Critérios Ofensores": Posicao (A) + Criterio (B) -> ranking fixo, direto do relatorio.
Aba "avaliacoes": Operador (D) + Nota Final (G) -> pega a PIOR nota de cada operador
e ranqueia os 4 operadores com a pior nota (agentes ofensores).
"""

import openpyxl
import os
import re
import glob

# ═══════════════════════════════════════════════════════════
# CONFIGURAÇÃO
# ═══════════════════════════════════════════════════════════

PASTA       = r'Arquivos\atualizaveis'
PREFIXO     = 'Relatorio Monitoria_Firjan'
INDEX_HTML  = r'index.html'

ABA_CRITERIOS = 'Critérios Ofensores'
ABA_AVAL      = 'avaliacoes'

COL_OPERADOR = 3   # D
COL_NOTA     = 6   # G

N_AGENTES = 4


# ═══════════════════════════════════════════════════════════
# FUNÇÕES
# ═══════════════════════════════════════════════════════════

def encontrar_arquivo(pasta, prefixo):
    padrao = os.path.join(pasta, f'{prefixo}*.xlsx')
    arquivos = [a for a in glob.glob(padrao) if not os.path.basename(a).startswith('~$')]
    if not arquivos:
        raise FileNotFoundError(f'[ERRO] Arquivo nao encontrado: {prefixo}*.xlsx em {pasta}')
    return sorted(arquivos)[-1]


def calcular(caminho):
    print(f'  Lendo: {os.path.basename(caminho)}')
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)

    # --- Critérios Ofensores (ranking fixo do relatorio) ---
    criterios = []
    if ABA_CRITERIOS in wb.sheetnames:
        ws = wb[ABA_CRITERIOS]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            if not row or not row[0] or not row[1]:
                continue
            pos = str(row[0]).strip()
            crit = str(row[1]).strip()
            criterios.append((pos, crit))
    else:
        print(f'  [AVISO] Aba "{ABA_CRITERIOS}" nao encontrada.')

    # --- Agentes Ofensores (pior nota por operador, top N) ---
    agentes = []
    if ABA_AVAL in wb.sheetnames:
        ws = wb[ABA_AVAL]
        pior_por_op = {}
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            if not any(c is not None and str(c).strip() for c in row):
                continue
            op = row[COL_OPERADOR] if len(row) > COL_OPERADOR else None
            nota = row[COL_NOTA] if len(row) > COL_NOTA else None
            if op is None or nota is None:
                continue
            op = str(op).strip()
            if op not in pior_por_op or nota < pior_por_op[op]:
                pior_por_op[op] = nota
        ranking = sorted(pior_por_op.items(), key=lambda x: x[1])[:N_AGENTES]
        agentes = [(f'{i+1}º', nome, nota) for i, (nome, nota) in enumerate(ranking)]
    else:
        print(f'  [AVISO] Aba "{ABA_AVAL}" nao encontrada.')

    wb.close()

    print(f'  Criterios Ofensores: {len(criterios)}')
    print(f'  Agentes Ofensores: {[f"{n} ({no})" for _, n, no in agentes]}')

    return {'criterios': criterios, 'agentes': agentes}


def js_str(v):
    return "'" + str(v).replace(chr(39), chr(92) + chr(39)) + "'"


def gerar_bloco(k):
    crit_js = ',\n    '.join(
        f"{{pos:{js_str(pos)}, criterio:{js_str(crit)}}}" for pos, crit in k['criterios']
    )
    ag_js = ',\n    '.join(
        f"{{pos:{js_str(pos)}, nome:{js_str(nome)}, nota:{nota}}}" for pos, nome, nota in k['agentes']
    )
    return f"""  /* QUALIDADE_START */
  const CRITERIOS_OFENSORES = [
    {crit_js}
  ];
  const AGENTES_OFENSORES = [
    {ag_js}
  ];
  /* QUALIDADE_END */"""


def atualizar_html(index_path, bloco):
    with open(index_path, 'r', encoding='utf-8') as f:
        conteudo = f.read()
    padrao = r'/\* QUALIDADE_START \*/.*?/\* QUALIDADE_END \*/'
    if not re.search(padrao, conteudo, re.DOTALL):
        raise ValueError('[ERRO] Marcadores QUALIDADE nao encontrados no index.html.')
    conteudo = re.sub(padrao, bloco, conteudo, flags=re.DOTALL)
    with open(index_path, 'w', encoding='utf-8') as f:
        f.write(conteudo)


def main():
    print()
    print('=' * 50)
    print('  ATUALIZADOR — Qualidade (Monitoria)')
    print('=' * 50)
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    try:
        caminho = encontrar_arquivo(PASTA, PREFIXO)
        k = calcular(caminho)
        bloco = gerar_bloco(k)
        atualizar_html(INDEX_HTML, bloco)
        print()
        print('=' * 50)
        print('  CONCLUIDO! index.html atualizado.')
        print('  Rode publicar.bat para enviar ao GitHub.')
        print('=' * 50)
        print()
    except Exception as e:
        print(f'\n[ERRO] {e}')
        import traceback; traceback.print_exc()


if __name__ == '__main__':
    main()
