# -*- coding: utf-8 -*-
"""
Atualizador Dashboard Firjan - Modulo Promocao Saude
Le a planilha do Google Sheets (export CSV) e injeta os dados brutos no index.html.
Os filtros (Data, Acao, Especialidade) e KPIs sao recalculados no navegador.

Colunas usadas (0-based):
  A(0)  = Data            -> so conta linha com este campo preenchido
  F(5)  = Empresa         -> eixo X do grafico "Intencoes por Empresa"
  I(8)  = Acao            -> dropdown de filtro
  J(9)  = Especialidade   -> dropdown de filtro
  L(11) = Agendamento     -> "Sim" => Agendada
  O(14) = Cancelamento    -> "Sim" => Cancelada
  S(18) = Pendencia       -> "Sim" => Pendente
"""

import urllib.request
import csv
import io
import re
import os
from datetime import datetime, timedelta

import openpyxl

# ═══════════════════════════════════════════════════════════
# CONFIGURAÇÃO
# ═══════════════════════════════════════════════════════════

SHEET_ID  = '1bsovvrAr1a5vi9Ny4kujga732cfJp4zpRNTYphpxJos'
GID       = '0'
CSV_URL   = f'https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=csv&gid={GID}'

# GID da aba AUX (abrir planilha → clicar na aba AUX → copiar o número após "gid=" na URL)
AUX_GID   = '22905422'
AUX_URL   = f'https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=csv&gid={AUX_GID}' if AUX_GID else ''

INDEX_HTML = r'index.html'

# --- Base Confirmacoes (xlsx local) ---
CONF_PATH = r'Arquivos\atualizaveis\Base Confirmações.xlsx'
CONF_ABA_ODONTO = 'Odonto-Massagem-Fisio'
CONF_ABA_SAUDE  = 'Saude'
# Odonto: A=data encerramento, H=especialidade (acao)
CONF_OD_DATA = 0   # A
CONF_OD_ESP  = 7   # H
# Saude: C=data, G=Prof.Agenda (medico); acao fixa = "Saude Digital"
CONF_SA_DATA = 2   # C
CONF_SA_PROF = 6   # G
CONF_SAUDE_ACAO = 'Saúde Digital'

COL_DATA       = 0   # A
COL_EMP        = 5   # F
COL_ACAO       = 8   # I
COL_PROF       = 10  # K  Profissional (Nome do Medico)
COL_AGEND      = 11  # L
COL_DATA_AGEND = 12  # M  Data de agendamento
COL_CANC       = 14  # O
COL_PEND       = 18  # S

# ═══════════════════════════════════════════════════════════
# FUNÇÕES
# ═══════════════════════════════════════════════════════════

def baixar_aux(url):
    """Baixa a aba AUX e retorna {norm_nome: capacidade_diaria}."""
    if not url:
        return {}
    req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
    data = urllib.request.urlopen(req, timeout=40).read().decode('utf-8', errors='replace')
    rows = list(csv.reader(io.StringIO(data)))
    aux = {}
    for r in rows[1:]:
        if len(r) < 2 or not r[0].strip() or not r[1].strip():
            continue
        try:
            cap = int(str(r[1]).strip().replace(',', '').replace('.', ''))
        except ValueError:
            continue
        aux[norm_prof(r[0].strip())] = cap
    print(f'  AUX: {len(aux)} medicos com capacidade configurada.')
    return aux


def baixar_csv(url):
    print(f'  Baixando planilha do Google Sheets...')
    req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
    data = urllib.request.urlopen(req, timeout=40).read().decode('utf-8', errors='replace')
    rows = list(csv.reader(io.StringIO(data)))
    print(f'  {len(rows)} linhas lidas (com cabecalho).')
    return rows


def is_sim(v):
    return str(v).strip().lower() == 'sim'


def parse_data(v):
    """dd/mm/yyyy -> int yyyymmdd. Retorna 0 se nao parsear."""
    s = str(v).strip()
    m = re.match(r'(\d{1,2})/(\d{1,2})/(\d{2,4})', s)
    if not m:
        return 0
    d, mo, y = m.groups()
    y = int(y)
    if y < 100:
        y += 2000
    return y * 10000 + int(mo) * 100 + int(d)


def cel(row, idx):
    return row[idx].strip() if len(row) > idx else ''


def conf_to_dt(v):
    """Converte celula de data (datetime ou dd/mm/aaaa) em datetime. None se falhar."""
    if isinstance(v, datetime):
        return v
    s = str(v).strip() if v is not None else ''
    m = re.match(r'(\d{1,2})/(\d{1,2})/(\d{2,4})', s)
    if m:
        d, mo, y = m.groups()
        y = int(y) + (2000 if int(y) < 100 else 0)
        try:
            return datetime(y, int(mo), int(d))
        except ValueError:
            return None
    return None


def conf_ajustar_data(dt):
    """Data da confirmacao -> data prevista da consulta.
    Seg-Qui: +1 dia | Sex: +3 dias | resto: +1 dia. Retorna int yyyymmdd."""
    wd = dt.weekday()  # Mon=0 ... Sun=6
    delta = 3 if wd == 4 else 1
    nd = dt + timedelta(days=delta)
    return nd.year * 10000 + nd.month * 100 + nd.day


def conf_titulo(s):
    """Cada palavra com a primeira letra maiuscula (ANA CAROLINE -> Ana Caroline)."""
    return ' '.join(w[:1].upper() + w[1:].lower() for w in str(s).split())


def conf_primeiro_nome(nome):
    """Primeiro nome; se for 'Ana', inclui o segundo nome tambem. Normaliza para Title Case."""
    parts = str(nome).strip().split()
    if not parts:
        return ''
    if parts[0].lower() == 'ana' and len(parts) > 1:
        sel = parts[0] + ' ' + parts[1]
    else:
        sel = parts[0]
    return conf_titulo(sel)


def ler_confirmacoes(caminho, aco_list, aco_idx, prof_list, prof_idx, get_idx):
    """Le a Base Confirmacoes e retorna conf_rows = [[dt_ajustada, acao_idx, med_idx], ...].
    acao/medico sao mapeados nas MESMAS listas dos filtros da Saude."""
    conf_rows = []
    if not os.path.exists(caminho):
        print(f'  [AVISO] Base Confirmacoes nao encontrada: {caminho}')
        return conf_rows
    wbc = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    n_od = n_sa = 0
    # --- Odonto-Massagem-Fisio: data col A, acao = especialidade col H ---
    if CONF_ABA_ODONTO in wbc.sheetnames:
        ws = wbc[CONF_ABA_ODONTO]
        for i, r in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            dv = conf_to_dt(r[CONF_OD_DATA] if len(r) > CONF_OD_DATA else None)
            if not dv:
                continue
            dt = conf_ajustar_data(dv)
            esp = str(r[CONF_OD_ESP]).strip() if len(r) > CONF_OD_ESP and r[CONF_OD_ESP] else ''
            ai = get_idx(esp, aco_list, aco_idx) if esp else -1
            conf_rows.append([dt, ai, -1])
            n_od += 1
    else:
        print(f'  [AVISO] Aba "{CONF_ABA_ODONTO}" nao encontrada.')
    # --- Saude: data col C, acao fixa "Saude Digital", medico col G ---
    if CONF_ABA_SAUDE in wbc.sheetnames:
        from collections import Counter
        ws = wbc[CONF_ABA_SAUDE]
        ai_sa = get_idx(CONF_SAUDE_ACAO, aco_list, aco_idx)
        # 1) coleta as linhas (dt, nome_medico)
        sa_pre = []
        formas = {}   # chave sem acento -> Counter de formas exibidas
        for i, r in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            dv = conf_to_dt(r[CONF_SA_DATA] if len(r) > CONF_SA_DATA else None)
            if not dv:
                continue
            dt = conf_ajustar_data(dv)
            med = conf_primeiro_nome(r[CONF_SA_PROF]) if len(r) > CONF_SA_PROF and r[CONF_SA_PROF] else ''
            sa_pre.append((dt, med))
            if med:
                formas.setdefault(norm_prof(med), Counter())[med] += 1
        # 2) nome canonico (forma mais frequente) por chave sem acento
        canon_med = {k: c.most_common(1)[0][0] for k, c in formas.items()}
        # 3) monta as linhas usando o nome canonico
        for dt, med in sa_pre:
            if med:
                med = canon_med.get(norm_prof(med), med)
                si = get_idx(med, prof_list, prof_idx)
            else:
                si = -1
            conf_rows.append([dt, ai_sa, si])
            n_sa += 1
    else:
        print(f'  [AVISO] Aba "{CONF_ABA_SAUDE}" nao encontrada.')
    wbc.close()
    print(f'  Confirmacoes: Odonto/Massagem/Fisio={n_od} | Saude={n_sa} | Total={len(conf_rows)}')
    return conf_rows


def base_prof(nome):
    """Remove '(odonto)/(saude)/...' do nome do profissional."""
    return re.sub(r'\s*\(.*?\)', '', nome).strip()


def norm_prof(nome):
    """Chave de agrupamento: sem parenteses, sem acento, minuscula."""
    import unicodedata
    b = base_prof(nome).lower()
    b = unicodedata.normalize('NFKD', b)
    return ''.join(c for c in b if not unicodedata.combining(c))


def unificar_medicos(prof_list, data_rows, conf_rows):
    """Funde medicos iguais (mesmo nome ignorando acento/caixa) entre as duas fontes.
    Escolhe a forma mais usada (desempate: com acento, depois mais longa).
    Remapeia os indices em data_rows[3] e conf_rows[2]. Retorna nova prof_list."""
    def has_accent(s):
        return any(ord(c) > 127 for c in s)
    uso = {}
    for r in data_rows:
        uso[r[3]] = uso.get(r[3], 0) + 1
    for r in conf_rows:
        uso[r[2]] = uso.get(r[2], 0) + 1
    grupos = {}
    for idx, nome in enumerate(prof_list):
        k = norm_prof(nome) if nome else ''
        grupos.setdefault(k, []).append(idx)
    new_list, remap = [], {}
    for k, idxs in grupos.items():
        best = sorted(idxs, key=lambda i: (uso.get(i, 0), has_accent(prof_list[i]), len(prof_list[i])), reverse=True)[0]
        novo = len(new_list)
        new_list.append(prof_list[best])
        for i in idxs:
            remap[i] = novo
    for r in data_rows:
        r[3] = remap.get(r[3], r[3])
    for r in conf_rows:
        r[2] = remap.get(r[2], r[2])
    return new_list


def processar(rows, aux_cap=None):
    from collections import Counter
    if aux_cap is None:
        aux_cap = {}
    body = rows[1:]
    emp_list, aco_list, prof_list = [], [], []
    emp_idx, aco_idx, prof_idx = {}, {}, {}

    def get_idx(val, lst, mp):
        if val not in mp:
            mp[val] = len(lst)
            lst.append(val)
        return mp[val]

    # Pre-scan: nome canonico do profissional por chave normalizada (forma mais frequente)
    formas = {}
    for r in body:
        if not cel(r, COL_DATA):
            continue
        p = cel(r, COL_PROF)
        if not p:
            continue
        formas.setdefault(norm_prof(p), Counter())[base_prof(p)] += 1
    canon = {k: c.most_common(1)[0][0] for k, c in formas.items()}

    data_rows = []
    total = 0
    for r in body:
        if not any(c.strip() for c in r):
            continue
        if not cel(r, COL_DATA):   # so conta linha com Data (col A)
            continue
        total += 1
        dt  = parse_data(cel(r, COL_DATA))
        ei  = get_idx(cel(r, COL_EMP) or 'Firjan', emp_list, emp_idx)   # sem empresa = Firjan
        ai  = get_idx(cel(r, COL_ACAO), aco_list, aco_idx)
        praw = cel(r, COL_PROF)
        prof = canon.get(norm_prof(praw), base_prof(praw)) if praw else ''
        si  = get_idx(prof, prof_list, prof_idx)
        ag      = 1 if is_sim(cel(r, COL_AGEND)) else 0
        ca      = 1 if is_sim(cel(r, COL_CANC))  else 0
        pe      = 1 if is_sim(cel(r, COL_PEND))  else 0
        dt_ag   = parse_data(cel(r, COL_DATA_AGEND))
        data_rows.append([dt, ei, ai, si, ag, ca, pe, dt_ag])

    # Base Confirmacoes (mapeia acao/medico nas mesmas listas)
    conf_rows = ler_confirmacoes(CONF_PATH, aco_list, aco_idx, prof_list, prof_idx, get_idx)

    # Unifica medicos iguais entre as duas fontes (acento/caixa)
    prof_list = unificar_medicos(prof_list, data_rows, conf_rows)

    # Dropdowns nao mostram valor vazio
    aco_drop  = [a for a in aco_list if a]
    prof_drop = [p for p in prof_list if p]

    # Montar SAUDE_AUX: [[idx_medico, capacidade_diaria], ...]
    saude_aux = []
    for norm_nome, cap in aux_cap.items():
        for idx, nome in enumerate(prof_list):
            if norm_prof(nome) == norm_nome:
                saude_aux.append([idx, cap])
                break

    print(f'  Interacoes (Data preenchida): {total}')
    print(f'  Agendadas (L=Sim): {sum(x[4] for x in data_rows)}')
    print(f'  Canceladas (O=Sim): {sum(x[5] for x in data_rows)}')
    print(f'  Pendentes (S=Sim): {sum(x[6] for x in data_rows)}')
    print(f'  Empresas: {len(emp_list)} | Acoes: {len(aco_drop)} | Medicos: {len(prof_drop)}')
    print(f'  AUX mapeados no SAUDE_MED: {len(saude_aux)}')

    return emp_list, aco_list, prof_list, data_rows, saude_aux, conf_rows


def gerar_bloco(emp_list, aco_list, prof_list, data_rows, saude_aux=None, conf_rows=None):
    def js_str(lst):
        return '[' + ','.join("'" + str(v).replace('\\', '\\\\').replace("'", "\\'") + "'" for v in lst) + ']'
    def js_rows(rows):
        return '[' + ','.join('[' + ','.join(str(c) for c in r) + ']' for r in rows) + ']'
    aux_js = js_rows(saude_aux) if saude_aux else '[]'
    conf_js = js_rows(conf_rows) if conf_rows else '[]'

    return (
        '/* SAUDE_DATA_START */\n'
        f'const SAUDE_EMP={js_str(emp_list)};\n'
        f'const SAUDE_ACO={js_str(aco_list)};\n'
        f'const SAUDE_MED={js_str(prof_list)};\n'
        f'const SAUDE_ROWS={js_rows(data_rows)};\n'
        f'const SAUDE_AUX={aux_js};\n'
        f'const CONF_ROWS={conf_js};\n'
        '/* SAUDE_DATA_END */'
    )


def atualizar_html(index_path, bloco):
    with open(index_path, 'r', encoding='utf-8') as f:
        conteudo = f.read()
    padrao = r'/\* SAUDE_DATA_START \*/.*?/\* SAUDE_DATA_END \*/'
    if not re.search(padrao, conteudo, re.DOTALL):
        raise ValueError('[ERRO] Marcadores SAUDE_DATA nao encontrados no index.html.')
    conteudo = re.sub(padrao, lambda m: bloco, conteudo, flags=re.DOTALL)  # lambda = substituicao literal
    with open(index_path, 'w', encoding='utf-8') as f:
        f.write(conteudo)


def carimbar_atualizacao(index_path):
    """Grava a data/hora da atualizacao no cabecalho do dashboard."""
    from datetime import datetime
    agora = datetime.now().strftime('%d/%m/%Y %H:%M')
    with open(index_path, 'r', encoding='utf-8') as f:
        c = f.read()
    c = re.sub(r'<!--LU-->.*?<!--/LU-->', lambda m: f'<!--LU-->{agora}<!--/LU-->', c, flags=re.DOTALL)
    with open(index_path, 'w', encoding='utf-8') as f:
        f.write(c)
    return agora


def main():
    print()
    print('=' * 50)
    print('  ATUALIZADOR - Promocao Saude')
    print('=' * 50)
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    try:
        print('\n[1/4] Lendo Google Sheets (Acompanhamento)...')
        rows = baixar_csv(CSV_URL)

        print('\n[2/4] Lendo aba AUX (capacidade médicos)...')
        aux_cap = baixar_aux(AUX_URL) if AUX_GID else {}
        if not AUX_GID:
            print('  [AVISO] AUX_GID nao configurado — Taxa de Ocupacao ficara zerada.')
            print('  Abra a planilha, clique na aba AUX e copie o numero apos "gid=" na URL.')

        print('\n[3/4] Processando dados...')
        emp_list, aco_list, esp_list, data_rows, saude_aux, conf_rows = processar(rows, aux_cap)

        print('\n[4/4] Atualizando index.html...')
        bloco = gerar_bloco(emp_list, aco_list, esp_list, data_rows, saude_aux, conf_rows)
        atualizar_html(INDEX_HTML, bloco)
        ts = carimbar_atualizacao(INDEX_HTML)
        print(f'  Atualizado em: {ts}')

        print()
        print('=' * 50)
        print('  CONCLUIDO! index.html atualizado.')
        print('  Rode publicar.bat para enviar ao GitHub.')
        print('=' * 50)
        print()
    except Exception as e:
        print(f'\n[ERRO] {e}')
        import traceback; traceback.print_exc()


if __name__ == '__main__':
    main()
    try:
        input('Pressione ENTER para fechar...')
    except EOFError:
        pass
