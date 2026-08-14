# -*- coding: utf-8 -*-
"""
Atualizador Dashboard Firjan - Campanha PotencializEE
Le a base de clientes (Google Sheets) + a base de ligacoes (Discagem) e
atualiza o bloco 'potencializee' no index.html.

Base de clientes -> card "Empresas na Base" (cada linha = 1 empresa) e
grafico de Motivos de Nao Interesse (coluna L).
Base de ligacoes (Discagem) -> Tentativas, Interessados, Decisor,
Conversao, Media, Distribuicao por Status e Evolucao Diaria.
"""

import csv
import glob
import io
import os
import re
import urllib.request
from collections import Counter, defaultdict
from datetime import datetime, timedelta

# ═══════════════════════════════════════════════════════════
# CONFIGURAÇÃO
# ═══════════════════════════════════════════════════════════

SHEET_ID   = '1_CHnyCehvqxriEY4ZCjhoW4JJAU8XPCs'
GID        = '14657048'
INDEX_HTML = r'index.html'
DEPARA_PATH = r'Arquivos\bases_apoio\tab_de-para.xlsx'
PASTA_DISCAGEM = r'Arquivos\nao_atualizaveis\Ativo_Potencializee'

COL_MOTIVO = 11  # L - Motivo do não interesse (base de clientes)

STATUS_MAP = {}
LABELS_NAO_DECISOR = {'Telefonia', 'Tentativa', 'Engano', 'Alo'}
LABEL_INTERESSADO  = 'Interessado'

# Tabulacoes brutas que contam como "Interessado" (nao mexe no de-para para nao afetar outras campanhas)
INTERESSADO_LINK_LABEL = 'Interessado / Link Enviado'
INTERESSADO_RAW_LABELS = ['Interessado / Link Enviado', 'INTERESSADO']

# Grafico "Distribuicao" (por status) so considera as tabulacoes CPC
# 'Interessado / Link Enviado' e 'INTERESSADO' sao unificadas no rotulo 'Interessado / Link Enviado'
CPC_LABELS = ['INFORMAÇÕES POR E-MAIL', 'Retornar', 'NAO INTERESSADO',
              'Interessado / Link Enviado', 'INTERESSADO']
CPC_LABEL_OVERRIDE = {
    'Interessado / Link Enviado': INTERESSADO_LINK_LABEL,
    'INTERESSADO': INTERESSADO_LINK_LABEL,
}

# Grafico "Evolucao Diaria" (tentativas) so considera as tabulacoes de insucesso
INSUCESSO_LABELS = ['Falhou', 'Fora de Area / Cx de Mensagens', 'Ligação Muda',
                     'Não Atendeu', 'Ocupado', 'Tel Não Atende / Ocupado',
                     'Atendido', 'Engano', 'Cliente Desligou']


def _norm_status(s):
    import unicodedata
    b = unicodedata.normalize('NFKD', str(s).strip().lower())
    return ''.join(c for c in b if not unicodedata.combining(c))


CPC_MAP = {_norm_status(s): CPC_LABEL_OVERRIDE.get(s, s) for s in CPC_LABELS}
INSUCESSO_MAP = {_norm_status(s): s for s in INSUCESSO_LABELS}
INTERESSADO_RAW_NORM = {_norm_status(s) for s in INTERESSADO_RAW_LABELS}

MES_PT = {1:'Jan',2:'Fev',3:'Mar',4:'Abr',5:'Mai',6:'Jun',
          7:'Jul',8:'Ago',9:'Set',10:'Out',11:'Nov',12:'Dez'}


# ═══════════════════════════════════════════════════════════
# FUNÇÕES AUXILIARES
# ═══════════════════════════════════════════════════════════

def fmt_num(n):
    return f"{n:,}".replace(",", ".")


def fmt_pct(n):
    return f"{n:.2f}%".replace(".", ",")


def fmt_dec(n):
    return f"{n:.2f}".replace(".", ",")


def norm_tel(t):
    return re.sub(r'\D', '', str(t)) if t else ''


def to_datetime(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, (int, float)) and val > 0:
        try:
            return datetime(1899, 12, 30) + timedelta(days=float(val))
        except Exception:
            return None
    return None


def encontrar_unico_xlsx(pasta):
    arquivos = [a for a in glob.glob(os.path.join(pasta, '*.xlsx')) if not os.path.basename(a).startswith('~$')]
    if not arquivos:
        raise FileNotFoundError(f'[ERRO] Nenhum .xlsx encontrado em {pasta}')
    return sorted(arquivos)[-1]


def ler_depara(caminho):
    mapa = {}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
        ws = wb['status'] if 'status' in wb.sheetnames else wb.active
        first = True
        for row in ws.iter_rows(values_only=True):
            if first:
                first = False
                continue
            if row[0] and row[1]:
                mapa[str(row[0]).strip()] = str(row[1]).strip()
        wb.close()
        print(f'  De-para: {len(mapa)} entradas')
    except FileNotFoundError:
        print(f'  [AVISO] De-para nao encontrado: {caminho}')
    return mapa


def normalizar_status(raw):
    if raw is None:
        return None
    s = str(raw).strip()
    for k, v in STATUS_MAP.items():
        if s.upper() == k.upper():
            return v
    return s


# ═══════════════════════════════════════════════════════════
# BASE DE CLIENTES (Google Sheets)
# ═══════════════════════════════════════════════════════════

def baixar_csv(sheet_id, gid=None):
    url = f'https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv'
    if gid:
        url += f'&gid={gid}'
    resp = urllib.request.urlopen(url, timeout=60)
    conteudo = resp.read().decode('utf-8-sig')
    return list(csv.reader(io.StringIO(conteudo)))


def calcular_base():
    print('  Baixando base de clientes PotencializEE (Google Sheets)...')
    linhas = baixar_csv(SHEET_ID, GID)
    dados = linhas[1:]

    empresas = 0
    motivo_counter = Counter()
    motivo_display = {}  # chave normalizada (case-insensitive) -> rotulo de exibicao (primeira grafia encontrada)
    for r in dados:
        if not any(c.strip() for c in r):
            continue
        empresas += 1
        motivo = r[COL_MOTIVO].strip() if len(r) > COL_MOTIVO else ''
        if motivo:
            chave = motivo.lower()
            motivo_display.setdefault(chave, motivo)
            motivo_counter[chave] += 1

    motivo_items = motivo_counter.most_common()
    print(f'  Empresas na base: {empresas}')
    print(f'  Com motivo de nao interesse: {sum(motivo_counter.values())} | Motivos distintos: {len(motivo_items)}')

    return {
        'empresas':     empresas,
        'motivoLabels': [motivo_display[m[0]] for m in motivo_items],
        'motivoData':   [m[1] for m in motivo_items],
    }


# ═══════════════════════════════════════════════════════════
# BASE DE LIGAÇÕES (Discagem)
# ═══════════════════════════════════════════════════════════

def calcular_discagem(caminho):
    """Le o Discagem_Potencializee.xlsx.
    Cada linha com DATA (col A) = 1 tentativa de ligacao.
    Status: col L (STATUS_NEGOCIO); se vazia, usa col K (STATUS).
    Telefone: TIPO=Discador -> Origem (H); TIPO=Sainte -> Destino (I).
    """
    import openpyxl
    print(f'  Lendo Discagem: {os.path.basename(caminho)}')
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    aba = next((s for s in wb.sheetnames if s.startswith('chamadas_')), wb.sheetnames[0])
    ws = wb[aba]

    data_idx = 0    # A DATA
    orig_idx = 7    # H ORIGEM
    dest_idx = 8    # I DESTINO
    tipo_idx = 9    # J TIPO
    st_idx   = 10   # K STATUS
    sn_idx   = 11   # L STATUS_NEGOCIO

    total_tent = 0
    status_counter = Counter()
    raw_por_label  = defaultdict(set)
    naosucesso_counter = Counter()
    raw_por_label_ns   = defaultdict(set)
    decisor_count = 0
    tel_interesse = set()
    data_min = data_max = None

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        dt = to_datetime(row[data_idx])
        if not dt:
            continue
        total_tent += 1
        if data_min is None or dt < data_min:
            data_min = dt
        if data_max is None or dt > data_max:
            data_max = dt

        tipo = str(row[tipo_idx]).strip().upper() if len(row) > tipo_idx and row[tipo_idx] else ''
        if tipo == 'DISCADOR':
            tel = norm_tel(row[orig_idx])
        elif tipo == 'SAINTE':
            tel = norm_tel(row[dest_idx])
        else:
            tel = norm_tel(row[orig_idx]) or norm_tel(row[dest_idx])

        sn_raw = row[sn_idx] if len(row) > sn_idx else None
        st_raw = row[st_idx] if len(row) > st_idx else None
        raw    = str(sn_raw).strip() if sn_raw else (str(st_raw).strip() if st_raw else '')

        raw_norm = _norm_status(raw) if raw else ''
        if raw_norm in CPC_MAP:
            canon = CPC_MAP[raw_norm]
            status_counter[canon] += 1
            raw_por_label[canon].add(raw)
            decisor_count += 1
        else:
            ns_label = raw if raw else 'Tentativas de Contato Sem Sucesso'
            naosucesso_counter[ns_label] += 1
            raw_por_label_ns[ns_label].add(raw if raw else '(vazio)')
        if raw_norm in INTERESSADO_RAW_NORM:
            tel_interesse.add(tel)

    wb.close()

    st_items = status_counter.most_common()
    st_tooltips = [', '.join(sorted(raw_por_label.get(s[0], set()))) for s in st_items]
    ns_items = naosucesso_counter.most_common()
    ns_tooltips = [', '.join(sorted(raw_por_label_ns.get(s[0], set()))) for s in ns_items]

    print(f'  Total Tentativas: {total_tent}')
    print(f'  Decisor: {decisor_count} | Interessados: {len(tel_interesse)}')
    print(f'  Status: {dict(st_items[:5])} ...')
    print(f'  Sem sucesso: {dict(ns_items[:5])} ...')

    return {
        'tentativas':    total_tent,
        'decisor':       decisor_count,
        'interessados':  len(tel_interesse),
        'statusLabels':  [s[0] for s in st_items],
        'statusData':    [s[1] for s in st_items],
        'statusTooltips': st_tooltips,
        'naosucessoLabels':   [s[0] for s in ns_items],
        'naosucessoData':     [s[1] for s in ns_items],
        'naosucessoTooltips': ns_tooltips,
        'dataMin': data_min,
        'dataMax': data_max,
    }


# ═══════════════════════════════════════════════════════════
# GERAÇÃO DO BLOCO JS
# ═══════════════════════════════════════════════════════════

def js_str(lst):
    return '[' + ','.join(f"'{str(v).replace(chr(39), chr(92)+chr(39))}'" for v in lst) + ']'


def js_num(lst):
    return '[' + ','.join(str(v) for v in lst) + ']'


def gerar_bloco(base, disc):
    empresas = base['empresas']
    tent     = disc['tentativas']
    decisor  = disc['decisor']
    interess = disc['interessados']
    taxa     = (interess / decisor * 100) if decisor > 0 else 0
    media    = (tent / empresas) if empresas > 0 else 0

    if disc['dataMin'] and disc['dataMax']:
        dmin, dmax = disc['dataMin'], disc['dataMax']
        periodo = (f"{dmin.day:02d}/{MES_PT[dmin.month]} — {dmax.day:02d}/{MES_PT[dmax.month]}")
    else:
        periodo = ''
    status_labels = disc['statusLabels'] or ['Sem dados']
    status_data   = disc['statusData'] or [0]
    ns_labels     = disc['naosucessoLabels'] or ['Sem dados']
    ns_data       = disc['naosucessoData'] or [0]

    return f"""  /* POTENCIALIZEE_START */
  potencializee: {{
    label: '— PotencializEE', desc: 'Campanha PotencializEE — dados filtrados', periodo: '{periodo}',
    empresas: '{fmt_num(empresas)}', empresasLabel: '🏢 Empresas na Base',
    mediaLabel: '🔁 Média Tentativas/Empresa', mediaSub: 'por empresa',
    tentativas: '{fmt_num(tent)}', interessados: '{fmt_num(interess)}', conversao: '{fmt_pct(taxa)}',
    decisor: '{fmt_num(decisor)}', decisorLabel: '👤 Status de Sucesso', decisorSub: 'Apenas PotencializEE', media: '{fmt_dec(media)}', trend: '',
    distTitle: 'Contatos de Sucesso',
    statusLabels: {js_str(status_labels)}, statusData: {js_num(status_data)}, statusColors:null,
    statusTooltips: {js_str(disc['statusTooltips'])},
    evoTitle: 'Tentativas de Contato Sem Sucesso', evoBar: true,
    naosucessoLabels: {js_str(ns_labels)}, naosucessoData: {js_num(ns_data)},
    naosucessoTooltips: {js_str(disc['naosucessoTooltips'])},
    evolucaoLabels: [], tentDia: [], convDia: [],
    showWpp: false,
    wppTitle: '', wppDesc: '',
    wppKpiLabels: [], wppListLabels: [], wppPieLabels: [],
    wppEnv:'-', wppResp:'-', wppTaxa:'-', wppSem:'-', wppInfo:'-', wppEmail:'-', wppPie:[0,1],
    distToggle: false,
    showMotivo: true,
    motivoLabels: {js_str(base['motivoLabels'])},
    motivoData: {js_num(base['motivoData'])}
  }},
  /* POTENCIALIZEE_END */"""


def atualizar_html(index_path, blocos):
    with open(index_path, 'r', encoding='utf-8') as f:
        conteudo = f.read()
    for marcador, novo_bloco in blocos.items():
        padrao = f'/\\* {marcador}_START \\*/.*?/\\* {marcador}_END \\*/'
        if not re.search(padrao, conteudo, re.DOTALL):
            raise ValueError(f'[ERRO] Marcadores {marcador} nao encontrados no index.html.')
        conteudo = re.sub(padrao, novo_bloco, conteudo, flags=re.DOTALL)
    with open(index_path, 'w', encoding='utf-8') as f:
        f.write(conteudo)


# ═══════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════

def main():
    print()
    print('=' * 50)
    print('  ATUALIZADOR — PotencializEE')
    print('=' * 50)
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    try:
        STATUS_MAP.update(ler_depara(DEPARA_PATH))
        base = calcular_base()
        disc = calcular_discagem(encontrar_unico_xlsx(PASTA_DISCAGEM))
        bloco = {'POTENCIALIZEE': gerar_bloco(base, disc)}
        atualizar_html(INDEX_HTML, bloco)
        print()
        print('=' * 50)
        print('  CONCLUIDO! index.html atualizado.')
        print('  Rode publicar.bat para enviar ao GitHub.')
        print('=' * 50)
        print()
    except Exception as e:
        print(f'\n[ERRO] {e}')
        import traceback; traceback.print_exc()


if __name__ == '__main__':
    main()
