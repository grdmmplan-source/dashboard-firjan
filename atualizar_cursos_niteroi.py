# -*- coding: utf-8 -*-
"""
Atualizador Dashboard Firjan - Campanha Cursos Técnicos Niterói
Mailing : cada linha = 1 inscrição (sem CNPJ)
Retorno : arquivo de chamadas (mesmo padrão Smart Factory)
WhatsApp: Planilha1 com colunas Data/Hora, Contato, Identificador, Status
"""

import openpyxl
import re
import glob
import os
from collections import Counter, defaultdict
from datetime import datetime, timedelta

PASTA_NITEROI = r'Arquivos\nao_atualizaveis\Ativo_Cursos_Técnicos_Unidade_Niterói'
ABA_RETORNO   = 'chamadas_15-06-2026_104745'
INDEX_HTML    = r'index.html'
DEPARA_PATH   = r'Arquivos\bases_apoio\tab_de-para.xlsx'

NOSSO_NUMERO       = '2120384382'
LABELS_NAO_DECISOR = {'Telefonia', 'Tentativa', 'Engano', 'Alo'}
LABEL_INTERESSADO  = 'Interessado'

STATUS_MAP = {}

MES_PT = {1:'Jan',2:'Fev',3:'Mar',4:'Abr',5:'Mai',6:'Jun',
          7:'Jul',8:'Ago',9:'Set',10:'Out',11:'Nov',12:'Dez'}

# ═══════════════════════════════════════════════════════════
# AUXILIARES
# ═══════════════════════════════════════════════════════════

def fmt_num(n):
    return f"{n:,}".replace(",", ".")

def fmt_pct(n):
    return f"{n:.2f}%".replace(".", ",")

def fmt_dec(n):
    return f"{n:.2f}".replace(".", ",")

def norm_tel(t):
    return re.sub(r'\D', '', str(t)) if t else ''

def to_datetime(val):
    if val is None: return None
    if isinstance(val, datetime): return val
    if isinstance(val, (int, float)) and val > 0:
        try: return datetime(1899, 12, 30) + timedelta(days=float(val))
        except: return None
    return None

def ler_depara(caminho):
    mapa = {}
    try:
        wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
        ws = wb['status'] if 'status' in wb.sheetnames else wb.active
        first = True
        for row in ws.iter_rows(values_only=True):
            if first: first = False; continue
            if row[0] and row[1]:
                mapa[str(row[0]).strip()] = str(row[1]).strip()
        wb.close()
        print(f'  De-para: {len(mapa)} entradas')
    except FileNotFoundError:
        print(f'  [AVISO] De-para nao encontrado: {caminho}')
    return mapa

def normalizar_status(raw):
    if raw is None: return None
    s = str(raw).strip()
    for k, v in STATUS_MAP.items():
        if s.upper() == k.upper(): return v
    return s

def encontrar_arquivo(pasta, prefixo):
    padrao = os.path.join(pasta, f'{prefixo}*.xlsx')
    arquivos = [a for a in glob.glob(padrao) if not os.path.basename(a).startswith('~$')]
    if not arquivos:
        raise FileNotFoundError(f'\n[ERRO] Arquivo nao encontrado: {prefixo}*.xlsx em {pasta}')
    return sorted(arquivos)[-1]

# ═══════════════════════════════════════════════════════════
# LEITURA
# ═══════════════════════════════════════════════════════════

def calcular_mailing_niteroi(caminho):
    """Mailing: apenas conta o total de linhas (cada linha = 1 inscrição).
    Status, tentativas e evolução vêm exclusivamente do Retorno."""
    print(f'  Lendo Mailing: {os.path.basename(caminho)}')
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb.active
    inscricoes = sum(1 for i, _ in enumerate(ws.iter_rows(values_only=True)) if i > 0)
    wb.close()
    print(f'    CPFs Inscritos: {inscricoes}')
    return {
        '_empresas':   inscricoes,
        '_inscricoes': inscricoes,
    }


def calcular_retorno_niteroi(caminho, aba):
    """Retorno: mesmo padrão do Smart Factory."""
    print(f'  Lendo Retorno: {os.path.basename(caminho)} | aba: {aba}')
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb[aba]
    next(ws.iter_rows(values_only=True))  # pula header
    wb_rows = list(ws.iter_rows(values_only=True, min_row=2))
    wb.close()

    data_idx = 0   # DATA
    orig_idx = 7   # ORIGEM
    dest_idx = 8   # DESTINO
    st_idx   = 10  # STATUS
    sn_idx   = 11  # STATUS_NEGOCIO

    total_tent     = 0
    status_counter = Counter()
    raw_por_label  = defaultdict(set)
    evolucao       = {}
    tel_decisor    = set()
    tel_interesse  = set()

    for row in wb_rows:
        dt = to_datetime(row[data_idx])
        if not dt: continue

        total_tent += 1

        orig_norm = norm_tel(row[orig_idx])
        tel = norm_tel(row[dest_idx]) if orig_norm == NOSSO_NUMERO else orig_norm

        sn_raw = row[sn_idx]
        st_raw = row[st_idx]
        raw    = str(sn_raw).strip() if sn_raw else (str(st_raw).strip() if st_raw else '')
        label  = normalizar_status(raw) if raw else None

        if label:
            status_counter[label] += 1
            if raw: raw_por_label[label].add(raw)

        if label and label not in LABELS_NAO_DECISOR:
            tel_decisor.add(tel)

        if label == LABEL_INTERESSADO:
            tel_interesse.add(tel)

        dk = f"{dt.day:02d}/{MES_PT[dt.month]}"
        if dk not in evolucao:
            evolucao[dk] = {'date': dt.date(), 'tent': 0, 'conv': 0}
        evolucao[dk]['tent'] += 1
        if label == LABEL_INTERESSADO:
            evolucao[dk]['conv'] += 1

    return {
        '_tentativas':    total_tent,
        '_statusCounter': status_counter,
        '_raw_por_label': dict(raw_por_label),
        '_evolucao':      evolucao,
        '_tel_decisor':   tel_decisor,
        '_tel_interesse': tel_interesse,
    }


def calcular_wpp_niteroi(caminho):
    """WhatsApp: Planilha1 — Status: Erro, Enviado, Entregue, Lido."""
    print(f'  Lendo WhatsApp: {os.path.basename(caminho)}')
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb.active
    status_counter = Counter()
    total = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0: continue
        total += 1
        st = row[3]
        if st: status_counter[str(st).strip()] += 1
    wb.close()

    entregues = status_counter.get('Entregue', 0) + status_counter.get('Lido', 0)
    lidos     = status_counter.get('Lido', 0)
    enviados  = status_counter.get('Enviado', 0)
    erros     = status_counter.get('Erro', 0)
    taxa      = (entregues / total * 100) if total > 0 else 0

    print(f'    Total: {total}  |  Entregues: {entregues}  |  Lidos: {lidos}  |  Enviado: {enviados}  |  Erro: {erros}')

    return {
        '_wpp_total':     total,
        '_wpp_entregues': entregues,
        '_wpp_lidos':     lidos,
        '_wpp_enviados':  enviados,
        '_wpp_erros':     erros,
        '_wpp_taxa':      taxa,
    }


def combinar_niteroi(m, r, w=None):
    """Mailing fornece só a contagem de inscrições.
    Status, tentativas e evolução vêm exclusivamente do Retorno."""
    inscricoes = m['_inscricoes']
    tent       = r['_tentativas']
    decisor    = len(r['_tel_decisor'])
    interesse  = len(r['_tel_interesse'])
    taxa  = (interesse / decisor * 100) if decisor > 0 else 0
    media = tent / inscricoes if inscricoes > 0 else 0

    st_items = r['_statusCounter'].most_common()
    rpl      = r.get('_raw_por_label', {})
    st_tooltips = [', '.join(sorted(rpl.get(s[0], set()))) for s in st_items]

    dias_ord = sorted(r['_evolucao'].items(), key=lambda x: x[1]['date'] or datetime(9999,12,31).date())

    show_wpp  = w is not None
    wpp_env   = fmt_num(w['_wpp_total'])     if w else '-'
    wpp_resp  = fmt_num(w['_wpp_entregues']) if w else '-'
    wpp_taxa  = fmt_pct(w['_wpp_taxa'])      if w else '-'
    wpp_sem   = fmt_num(w['_wpp_erros'])     if w else '-'
    wpp_info  = fmt_num(w['_wpp_lidos'])     if w else '-'
    wpp_email = fmt_num(w['_wpp_enviados'])  if w else '-'
    wpp_pie   = [w['_wpp_entregues'], w['_wpp_erros']] if w else [0, 1]

    return {
        # compatível com somar_campanhas
        '_empresas':      inscricoes,
        '_inscricoes':    inscricoes,
        '_tentativas':    tent,
        '_interessados':  interesse,
        '_decisor':       decisor,
        '_statusCounter': r['_statusCounter'],
        '_raw_por_label': rpl,
        '_evolucao':      r['_evolucao'],
        # formatados para o dashboard
        'inscricoes':    fmt_num(inscricoes),
        'tentativas':    fmt_num(tent),
        'interessados':  fmt_num(interesse),
        'conversao':     fmt_pct(taxa),
        'decisor':       fmt_num(decisor),
        'media':         fmt_dec(media),
        'statusLabels':   [s[0] for s in st_items],
        'statusData':     [s[1] for s in st_items],
        'statusTooltips': st_tooltips,
        'evoLabels':      [d[0] for d in dias_ord],
        'tentDia':        [d[1]['tent'] for d in dias_ord],
        'convDia':        [d[1]['conv'] for d in dias_ord],
        'showWpp':       show_wpp,
        'wppEnv':        wpp_env,
        'wppResp':       wpp_resp,
        'wppTaxa':       wpp_taxa,
        'wppSem':        wpp_sem,
        'wppInfo':       wpp_info,
        'wppEmail':      wpp_email,
        'wppPie':        wpp_pie,
    }


def calcular_kpis_niteroi():
    m = calcular_mailing_niteroi(encontrar_arquivo(PASTA_NITEROI, 'Mailing_Cursos'))
    r = calcular_retorno_niteroi(encontrar_arquivo(PASTA_NITEROI, 'Retorno_Ativo_Cursos'), ABA_RETORNO)
    try:
        w = calcular_wpp_niteroi(encontrar_arquivo(PASTA_NITEROI, 'Retorno_Whatsapp_Cursos'))
    except FileNotFoundError as e:
        print(f'  [AVISO] WhatsApp nao encontrado — secao sera ocultada. {e}')
        w = None
    return combinar_niteroi(m, r, w)

# ═══════════════════════════════════════════════════════════
# GERAÇÃO DO BLOCO JS
# ═══════════════════════════════════════════════════════════

def gerar_bloco_niteroi(k):
    def js_str(lst): return '[' + ','.join(f"'{str(v).replace(chr(39), chr(92)+chr(39))}'" for v in lst) + ']'
    def js_num(lst): return '[' + ','.join(str(v) for v in lst) + ']'

    show_wpp = 'true' if k.get('showWpp') else 'false'
    wpp_pie  = js_num(k.get('wppPie', [0, 1]))

    periodo = f"{k['evoLabels'][0]} — {k['evoLabels'][-1]}" if k['evoLabels'] else ''
    return f"""  /* NITEROI_START */
  niteroi: {{
    label: '— Cursos Técnicos Niterói', desc: 'Campanha Cursos Técnicos Niterói — dados filtrados', periodo: '{periodo}',
    empresas: '{k['inscricoes']}', empresasLabel: '📋 CPFs Inscritos',
    mediaLabel: '🔁 Média Tentativas/Inscrição', mediaSub: 'por inscrição',
    tentativas: '{k['tentativas']}', interessados: '{k['interessados']}', conversao: '{k['conversao']}',
    decisor: '{k['decisor']}', decisorSub: 'Apenas Cursos Técnicos Niterói', media: '{k['media']}', trend: '',
    statusLabels: {js_str(k['statusLabels'])},
    statusData: {js_num(k['statusData'])}, statusColors:null,
    statusTooltips: {js_str(k.get('statusTooltips', ['']*len(k['statusLabels'])))},
    evolucaoLabels: {js_str(k['evoLabels'])},
    tentDia: {js_num(k['tentDia'])}, convDia: {js_num(k['convDia'])},
    showWpp: {show_wpp},
    wppTitle: 'WhatsApp — Cursos Técnicos Niterói',
    wppDesc: 'Disparos massivos da campanha Cursos Técnicos Niterói',
    wppKpiLabels: ['📤 Total Disparado','✅ Entregues','📊 Taxa de Entrega','❌ Com Erro'],
    wppListLabels: ['Total Disparado','Lidos','Enviado','Com Erro'],
    wppPieLabels: ['Entregues','Com Erro'],
    wppEnv:'{k.get('wppEnv','-')}', wppResp:'{k.get('wppResp','-')}', wppTaxa:'{k.get('wppTaxa','-')}', wppSem:'{k.get('wppSem','-')}', wppInfo:'{k.get('wppInfo','-')}', wppEmail:'{k.get('wppEmail','-')}', wppPie:{wpp_pie}
  }},
  /* NITEROI_END */"""

# ═══════════════════════════════════════════════════════════
# ATUALIZAÇÃO DO INDEX.HTML
# ═══════════════════════════════════════════════════════════

def atualizar_html(index_path, blocos):
    with open(index_path, 'r', encoding='utf-8') as f:
        conteudo = f.read()
    for marcador, novo_bloco in blocos.items():
        padrao = f'/\\* {marcador}_START \\*/.*?/\\* {marcador}_END \\*/'
        if not re.search(padrao, conteudo, re.DOTALL):
            raise ValueError(f'[ERRO] Marcadores {marcador} nao encontrados no index.html.')
        conteudo = re.sub(padrao, novo_bloco, conteudo, flags=re.DOTALL)
    with open(index_path, 'w', encoding='utf-8') as f:
        f.write(conteudo)

# ═══════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════

def main():
    print()
    print('=' * 50)
    print('  ATUALIZADOR — Cursos Técnicos Niterói')
    print('=' * 50)

    os.chdir(os.path.dirname(os.path.abspath(__file__)))

    try:
        print('\n[0/3] Carregando de-para...')
        STATUS_MAP.update(ler_depara(DEPARA_PATH))

        print('\n[1/3] Calculando KPIs — Cursos Técnicos Niterói...')
        kn = calcular_kpis_niteroi()

        print(f'  CPFs Inscritos   : {kn["inscricoes"]}')
        print(f'  Total Tentativas : {kn["tentativas"]}')
        print(f'  Interessados     : {kn["interessados"]}')
        print(f'  Contatos Decisor : {kn["decisor"]}')
        print(f'  Taxa Conversão   : {kn["conversao"]}')
        print(f'  Média Tent/Insc  : {kn["media"]}')

        print('\n[2/3] Gerando bloco JavaScript...')
        blocos = {'NITEROI': gerar_bloco_niteroi(kn)}

        print('\n[3/3] Atualizando index.html...')
        atualizar_html(INDEX_HTML, blocos)

        print()
        print('=' * 50)
        print('  CONCLUIDO! index.html atualizado.')
        print('  Rode publicar.bat para enviar ao GitHub.')
        print('=' * 50)
        print()

    except Exception as e:
        print(f'\n[ERRO] {e}')
        import traceback; traceback.print_exc()

if __name__ == '__main__':
    import sys
    main()
    if '--no-pause' not in sys.argv:
        input('Pressione ENTER para fechar...')
