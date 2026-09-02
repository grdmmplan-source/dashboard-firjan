# -*- coding: utf-8 -*-
"""
Atualizador Dashboard Firjan - Ocupacao (Receptivo)
Le Base Ocupacao.xlsx e atualiza o bloco OCUPACAO no index.html.

Aba "Tempos Callflex" (canal Telefone/Voz):
  DATA_LOGIN (D) -> filtro de data
  TEMPO_TOTAL_LOGADO (G), TEMPO_TOTAL_DE_PAUSAS (N), TEMPO_EM_ATENDIMENTO (T)
  formato "HH:MM:SS" (ou "--" quando sem tempo -> tratado como 00:00:00)
  Ocupacao = soma(TEMPO_EM_ATENDIMENTO) / soma(TEMPO_TOTAL_LOGADO - TEMPO_TOTAL_DE_PAUSAS)

Aba "Tempos Sales" (demais canais digitais, agregados sem distincao por canal):
  Data (L) -> filtro de data
  Tipo_tempo (K): 'tempo_em_atendimento' ou 'tempo_total_de_pausas'
  Duracao do status (F) e Duracao de tempo ocioso (D), em segundos
  Ocupacao = soma(F-D onde tipo=tempo_em_atendimento) /
             (soma(F de todas as linhas) - soma(F onde tipo=tempo_total_de_pausas))
"""

import openpyxl
import os
import re
import glob
import datetime

PASTA      = r'Arquivos\atualizaveis'
PREFIXO    = 'Base Ocupação'
INDEX_HTML = r'index.html'

# --- Aba Tempos Callflex ---
CF_DATA   = 3    # D  DATA_LOGIN
CF_LOGADO = 6    # G  TEMPO_TOTAL_LOGADO
CF_PAUSAS = 13   # N  TEMPO_TOTAL_DE_PAUSAS
CF_ATEND  = 19   # T  TEMPO_EM_ATENDIMENTO

# --- Aba Tempos Sales ---
SL_DUR_STATUS = 5   # F  Duracao do status (segundos)
SL_DUR_OCIOSO = 3   # D  Duracao de tempo ocioso (segundos)
SL_TIPO       = 10  # K  Tipo_tempo
SL_DATA       = 11  # L  Data

TIPO_ATENDIMENTO = 'tempo_em_atendimento'
TIPO_PAUSAS      = 'tempo_total_de_pausas'


def encontrar_arquivo(pasta, prefixo):
    padrao = os.path.join(pasta, f'{prefixo}*.xlsx')
    arquivos = [a for a in glob.glob(padrao) if not os.path.basename(a).startswith('~$')]
    if not arquivos:
        raise FileNotFoundError(f'[ERRO] Arquivo nao encontrado: {prefixo}*.xlsx em {pasta}')
    return sorted(arquivos)[-1]


def hms_para_segundos(v):
    """Converte 'HH:MM:SS' (ou variantes tipo '-', '--', vazio) para segundos. '--'/vazio -> 0."""
    if v is None:
        return 0
    if isinstance(v, datetime.time):
        return v.hour * 3600 + v.minute * 60 + v.second
    if isinstance(v, datetime.timedelta):
        return int(v.total_seconds())
    s = str(v).strip()
    if not s or set(s) <= {'-'}:
        return 0
    m = re.match(r'^(\d+):(\d{2}):(\d{2})$', s)
    if m:
        h, mi, se = (int(x) for x in m.groups())
        return h * 3600 + mi * 60 + se
    m = re.match(r'^(\d{2}):(\d{2})$', s)
    if m:
        mi, se = (int(x) for x in m.groups())
        return mi * 60 + se
    return 0


def data_iso_para_int(v):
    """DATA_LOGIN vem como 'YYYY-MM-DD' (ou datetime)."""
    if v is None:
        return None
    if isinstance(v, (datetime.date, datetime.datetime)):
        return v.year * 10000 + v.month * 100 + v.day
    s = str(v).strip()
    m = re.match(r'^(\d{4})-(\d{2})-(\d{2})', s)
    if m:
        yy, mo, dd = (int(x) for x in m.groups())
        return yy * 10000 + mo * 100 + dd
    return None


def data_br_para_int(v):
    """Data (Tempos Sales) vem como 'DD/MM/YYYY' (ou datetime)."""
    if v is None:
        return None
    if isinstance(v, (datetime.date, datetime.datetime)):
        return v.year * 10000 + v.month * 100 + v.day
    s = str(v).strip()
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})', s)
    if m:
        dd, mo, yy = (int(x) for x in m.groups())
        return yy * 10000 + mo * 100 + dd
    return None


def calcular_callflex(caminho):
    print('  Lendo aba "Tempos Callflex"...')
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb['Tempos Callflex']
    linhas = []
    total_atend = total_logado = total_pausas = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if not row or row[CF_DATA] is None:
            continue
        dt = data_iso_para_int(row[CF_DATA])
        if dt is None:
            continue
        logado = hms_para_segundos(row[CF_LOGADO] if len(row) > CF_LOGADO else None)
        pausas = hms_para_segundos(row[CF_PAUSAS] if len(row) > CF_PAUSAS else None)
        atend  = hms_para_segundos(row[CF_ATEND] if len(row) > CF_ATEND else None)
        linhas.append([dt, atend, logado, pausas])
        total_atend += atend
        total_logado += logado
        total_pausas += pausas
    wb.close()
    den = total_logado - total_pausas
    pct = (total_atend / den * 100) if den > 0 else None
    print(f'  Callflex: {len(linhas)} linhas | Atendimento={total_atend}s | Logado-Pausas={den}s'
          + (f' | Ocupacao={pct:.2f}%' if pct is not None else ' | Ocupacao=--'))
    return linhas


def calcular_sales(caminho):
    print('  Lendo aba "Tempos Sales"...')
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb['Tempos Sales']
    linhas = []
    num = total_logado = pausas = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if not row or row[SL_DATA] is None:
            continue
        dt = data_br_para_int(row[SL_DATA])
        if dt is None:
            continue
        tipo = str(row[SL_TIPO]).strip().lower() if row[SL_TIPO] is not None else ''
        f = row[SL_DUR_STATUS] if len(row) > SL_DUR_STATUS and row[SL_DUR_STATUS] is not None else 0
        d = row[SL_DUR_OCIOSO] if len(row) > SL_DUR_OCIOSO and row[SL_DUR_OCIOSO] is not None else 0
        f = float(f) if isinstance(f, (int, float)) else 0
        d = float(d) if isinstance(d, (int, float)) else 0
        tipo_flag = 1 if tipo == TIPO_ATENDIMENTO else (2 if tipo == TIPO_PAUSAS else 0)
        linhas.append([dt, tipo_flag, f, d])
        total_logado += f
        if tipo == TIPO_ATENDIMENTO:
            num += (f - d)
        elif tipo == TIPO_PAUSAS:
            pausas += f
    wb.close()
    den = total_logado - pausas
    pct = (num / den * 100) if den > 0 else None
    print(f'  Sales: {len(linhas)} linhas | Numerador={num:.0f}s | Logado-Pausas={den:.0f}s'
          + (f' | Ocupacao={pct:.2f}%' if pct is not None else ' | Ocupacao=--'))
    return linhas


def js_rows(rows):
    partes = []
    for r in rows:
        cells = []
        for c in r:
            if c is None:
                cells.append('null')
            elif isinstance(c, float) and c.is_integer():
                cells.append(str(int(c)))
            else:
                cells.append(str(c))
        partes.append('[' + ','.join(cells) + ']')
    return '[' + ','.join(partes) + ']'


def gerar_bloco(callflex_rows, sales_rows):
    return f"""  /* OCUPACAO_START */
  const OCUP_CALLFLEX_ROWS = {js_rows(callflex_rows)};
  const OCUP_SALES_ROWS = {js_rows(sales_rows)};
  /* OCUPACAO_END */"""


def atualizar_html(index_path, bloco):
    with open(index_path, 'r', encoding='utf-8') as f:
        conteudo = f.read()
    padrao = r'/\* OCUPACAO_START \*/.*?/\* OCUPACAO_END \*/'
    if not re.search(padrao, conteudo, re.DOTALL):
        raise ValueError('[ERRO] Marcadores OCUPACAO nao encontrados no index.html.')
    conteudo = re.sub(padrao, lambda m: bloco, conteudo, flags=re.DOTALL)
    with open(index_path, 'w', encoding='utf-8') as f:
        f.write(conteudo)


def main():
    print()
    print('=' * 50)
    print('  ATUALIZADOR — Ocupação (Receptivo)')
    print('=' * 50)
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    try:
        caminho = encontrar_arquivo(PASTA, PREFIXO)
        print(f'  Lendo: {os.path.basename(caminho)}')
        callflex_rows = calcular_callflex(caminho)
        sales_rows = calcular_sales(caminho)
        bloco = gerar_bloco(callflex_rows, sales_rows)
        atualizar_html(INDEX_HTML, bloco)
        print()
        print('=' * 50)
        print('  CONCLUIDO! index.html atualizado.')
        print('  Rode publicar.bat para enviar ao GitHub.')
        print('=' * 50)
        print()
    except Exception as e:
        print(f'\n[ERRO] {e}')
        import traceback; traceback.print_exc()


if __name__ == '__main__':
    main()
