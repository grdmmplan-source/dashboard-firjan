# -*- coding: utf-8 -*-
"""
Atualizador Dashboard Firjan - Campanha Prospeccao IEL
Le a base de empresas (Google Sheets) + a base de ligacoes (Discagem) e
atualiza o bloco 'iel' no index.html.

Base de empresas -> card "Empresas na Base".
Base de ligacoes (Discagem) -> Tentativas, Interessados, Decisor,
Conversao, Media, Distribuicao por Status e Evolucao Diaria.
"""

import csv
import glob
import io
import os
import re
import urllib.request
from collections import Counter, defaultdict
from datetime import datetime, timedelta

# ═══════════════════════════════════════════════════════════
# CONFIGURAÇÃO
# ═══════════════════════════════════════════════════════════

SHEET_ID   = '10epaIyncKbZO9auh8QWnVm8f9VTCzLCY'
GID        = '1369157082'
INDEX_HTML = r'index.html'
DEPARA_PATH    = r'Arquivos\bases_apoio\tab_de-para.xlsx'
PASTA_DISCAGEM = r'Arquivos\nao_atualizaveis\Ativo_Prospecção IEL'

STATUS_MAP = {}
LABEL_INTERESSADO = 'Interessado'

MES_PT = {1:'Jan',2:'Fev',3:'Mar',4:'Abr',5:'Mai',6:'Jun',
          7:'Jul',8:'Ago',9:'Set',10:'Out',11:'Nov',12:'Dez'}


# ═══════════════════════════════════════════════════════════
# FUNÇÕES AUXILIARES
# ═══════════════════════════════════════════════════════════

def fmt_num(n):
    return f"{n:,}".replace(",", ".")


def fmt_pct(n):
    return f"{n:.2f}%".replace(".", ",")


def fmt_dec(n):
    return f"{n:.2f}".replace(".", ",")


def norm_tel(t):
    return re.sub(r'\D', '', str(t)) if t else ''


def to_datetime(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, (int, float)) and val > 0:
        try:
            return datetime(1899, 12, 30) + timedelta(days=float(val))
        except Exception:
            return None
    return None


def encontrar_unico_xlsx(pasta):
    arquivos = [a for a in glob.glob(os.path.join(pasta, '*.xlsx')) if not os.path.basename(a).startswith('~$')]
    if not arquivos:
        raise FileNotFoundError(f'[ERRO] Nenhum .xlsx encontrado em {pasta}')
    return sorted(arquivos)[-1]


def ler_depara(caminho):
    mapa = {}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
        ws = wb['status'] if 'status' in wb.sheetnames else wb.active
        first = True
        for row in ws.iter_rows(values_only=True):
            if first:
                first = False
                continue
            if row[0] and row[1]:
                mapa[str(row[0]).strip()] = str(row[1]).strip()
        wb.close()
        print(f'  De-para: {len(mapa)} entradas')
    except FileNotFoundError:
        print(f'  [AVISO] De-para nao encontrado: {caminho}')
    return mapa


def normalizar_status(raw):
    if raw is None:
        return None
    s = str(raw).strip()
    for k, v in STATUS_MAP.items():
        if s.upper() == k.upper():
            return v
    return s


def norm_txt(s):
    """Maiuscula, sem acento - para comparar tabulacoes brutas."""
    import unicodedata
    b = unicodedata.normalize('NFKD', str(s).strip().upper())
    return ''.join(c for c in b if not unicodedata.combining(c))


# Tabulacoes brutas (antes do de-para) usadas nos cards especificos da IEL
TAB_CONTATADAS  = {'INTERESSADO', 'NAO INTERESSADO', 'INFORMACOES POR E-MAIL', 'CLIENTE DESLIGOU'}
TAB_EFETIVOS    = {'INTERESSADO', 'INFORMACOES POR E-MAIL'}
TAB_DPMEN       = {'INFORMACOES POR E-MAIL'}


# ═══════════════════════════════════════════════════════════
# BASE DE EMPRESAS (Google Sheets)
# ═══════════════════════════════════════════════════════════

def baixar_csv(sheet_id, gid=None):
    url = f'https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv'
    if gid:
        url += f'&gid={gid}'
    resp = urllib.request.urlopen(url, timeout=60)
    conteudo = resp.read().decode('utf-8-sig')
    return list(csv.reader(io.StringIO(conteudo)))


COL_MOTIVO = 10  # K - Motivo do não interesse

def calcular_base():
    print('  Baixando base de empresas Prospecção IEL (Google Sheets)...')
    linhas = baixar_csv(SHEET_ID, GID)
    dados = linhas[1:]

    empresas = 0
    motivo_counter = Counter()
    for r in dados:
        if not any(c.strip() for c in r):
            continue
        empresas += 1
        motivo = r[COL_MOTIVO].strip() if len(r) > COL_MOTIVO else ''
        if motivo:
            motivo_counter[motivo] += 1

    motivo_items = motivo_counter.most_common()
    print(f'  Empresas na base: {empresas}')
    print(f'  Com motivo de nao interesse: {sum(motivo_counter.values())} | Motivos distintos: {len(motivo_items)}')

    return {
        'empresas': empresas,
        'motivoLabels': [m[0] for m in motivo_items],
        'motivoData':   [m[1] for m in motivo_items],
    }


# ═══════════════════════════════════════════════════════════
# BASE DE LIGAÇÕES (Discagem)
# ═══════════════════════════════════════════════════════════

def calcular_discagem(caminho):
    """Le o Discagem_Ativo_Prospecção_IEL.xlsx.
    Cada linha com DATA (col A) = 1 tentativa de ligacao.
    Status: col L (STATUS_NEGOCIO); se vazia, usa col K (STATUS).
    """
    import openpyxl
    print(f'  Lendo Discagem: {os.path.basename(caminho)}')
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    aba = next((s for s in wb.sheetnames if s.startswith('chamadas_')), wb.sheetnames[0])
    ws = wb[aba]

    data_idx = 0    # A DATA
    doc_idx  = 4    # E DOC (CNPJ)
    orig_idx = 7    # H ORIGEM
    dest_idx = 8    # I DESTINO
    st_idx   = 10   # K STATUS
    sn_idx   = 11   # L STATUS_NEGOCIO

    total_tent = 0
    status_counter = Counter()
    raw_por_label  = defaultdict(set)
    evolucao = {}
    tel_interesse    = set()
    emp_contatadas   = set()  # empresas: INTERESSADO, NAO INTERESSADO, INFORMACOES POR E-MAIL, CLIENTE DESLIGOU
    emp_efetivos     = set()  # empresas: INTERESSADO, INFORMACOES POR E-MAIL
    emp_dpmen        = set()  # empresas: INFORMACOES POR E-MAIL

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        dt = to_datetime(row[data_idx])
        if not dt:
            continue
        total_tent += 1

        tel = norm_tel(row[orig_idx]) or norm_tel(row[dest_idx])
        doc = row[doc_idx] if len(row) > doc_idx else None
        empresa_id = doc if doc else tel

        sn_raw = row[sn_idx] if len(row) > sn_idx else None
        st_raw = row[st_idx] if len(row) > st_idx else None
        raw    = str(sn_raw).strip() if sn_raw else (str(st_raw).strip() if st_raw else '')
        label  = normalizar_status(raw) if raw else None
        rawn   = norm_txt(raw) if raw else ''

        if label:
            status_counter[label] += 1
            if raw:
                raw_por_label[label].add(raw)

        if rawn in TAB_CONTATADAS:
            emp_contatadas.add(empresa_id)
        if rawn in TAB_EFETIVOS:
            emp_efetivos.add(empresa_id)
        if rawn in TAB_DPMEN:
            emp_dpmen.add(empresa_id)
        if label == LABEL_INTERESSADO:
            tel_interesse.add(tel)

        dk = f"{dt.day:02d}/{MES_PT[dt.month]}"
        if dk not in evolucao:
            evolucao[dk] = {'date': dt.date(), 'tent': 0, 'conv': 0}
        evolucao[dk]['tent'] += 1
        if label == LABEL_INTERESSADO:
            evolucao[dk]['conv'] += 1

    wb.close()

    dias_ord = sorted(evolucao.items(), key=lambda x: x[1]['date'])
    st_items = status_counter.most_common()
    st_tooltips = [', '.join(sorted(raw_por_label.get(s[0], set()))) for s in st_items]

    print(f'  Total Tentativas: {total_tent}')
    print(f'  Empresas contatadas: {len(emp_contatadas)} | Interessados: {len(tel_interesse)}')
    print(f'  Contatos efetivos: {len(emp_efetivos)} | Leads DPMEN: {len(emp_dpmen)}')
    print(f'  Status: {dict(st_items[:5])} ...')

    return {
        'tentativas':    total_tent,
        'decisor':       len(emp_contatadas),
        'interessados':  len(tel_interesse),
        'efetivos':      len(emp_efetivos),
        'dpmen':         len(emp_dpmen),
        'statusLabels':  [s[0] for s in st_items],
        'statusData':    [s[1] for s in st_items],
        'statusTooltips': st_tooltips,
        'evoLabels':     [d[0] for d in dias_ord],
        'tentDia':       [d[1]['tent'] for d in dias_ord],
        'convDia':       [d[1]['conv'] for d in dias_ord],
    }


# ═══════════════════════════════════════════════════════════
# GERAÇÃO DO BLOCO JS
# ═══════════════════════════════════════════════════════════

def js_str(lst):
    return '[' + ','.join(f"'{str(v).replace(chr(39), chr(92)+chr(39))}'" for v in lst) + ']'


def js_num(lst):
    return '[' + ','.join(str(v) for v in lst) + ']'


def gerar_bloco(base, disc):
    empresas = base['empresas']
    tent     = disc['tentativas']
    decisor  = disc['decisor']       # Empresas contatadas
    interess = disc['interessados']
    efetivos = disc['efetivos']
    dpmen    = disc['dpmen']
    taxa     = (interess / decisor * 100) if decisor > 0 else 0
    media    = (tent / empresas) if empresas > 0 else 0
    taxa_interesse = (interess / decisor * 100) if decisor > 0 else 0

    periodo = f"{disc['evoLabels'][0]} — {disc['evoLabels'][-1]}" if disc['evoLabels'] else ''
    status_labels = disc['statusLabels'] or ['Sem dados']
    status_data   = disc['statusData'] or [0]
    evo_labels    = disc['evoLabels'] or ['--']
    tent_dia      = disc['tentDia'] or [0]
    conv_dia      = disc['convDia'] or [0]

    return f"""  /* IEL_START */
  iel: {{
    label: '— Prospecção IEL', desc: 'Campanha Prospecção IEL — dados filtrados', periodo: '{periodo}',
    empresas: '{fmt_num(empresas)}', empresasLabel: '🏢 Empresas na Base',
    mediaLabel: '🔁 Média Tentativas/Empresa', mediaSub: 'por empresa',
    tentativas: '{fmt_num(tent)}', interessados: '{fmt_num(interess)}', conversao: '{fmt_pct(taxa)}',
    decisor: '{fmt_num(decisor)}', decisorLabel: '📋 Empresas Contatadas', decisorSub: 'Apenas Prospecção IEL', media: '{fmt_dec(media)}', trend: '',
    statusLabels: {js_str(status_labels)}, statusData: {js_num(status_data)}, statusColors:null,
    statusTooltips: {js_str(disc['statusTooltips'])},
    evolucaoLabels: {js_str(evo_labels)},
    tentDia: {js_num(tent_dia)}, convDia: {js_num(conv_dia)},
    showWpp: false,
    wppTitle: '', wppDesc: '',
    wppKpiLabels: [], wppListLabels: [], wppPieLabels: [],
    wppEnv:'-', wppResp:'-', wppTaxa:'-', wppSem:'-', wppInfo:'-', wppEmail:'-', wppPie:[0,1],
    hideConvMedia: true,
    showIelExtra: true,
    ielEfetivos: '{fmt_num(efetivos)}',
    ielTaxaInteresse: '{fmt_pct(taxa_interesse)}',
    ielDpmen: '{fmt_num(dpmen)}',
    ielInscritas: '-',
    ielTaxaConv: '-',
    showMotivo: true,
    motivoLabels: {js_str(base['motivoLabels'])},
    motivoData: {js_num(base['motivoData'])}
  }},
  /* IEL_END */"""


def atualizar_html(index_path, blocos):
    with open(index_path, 'r', encoding='utf-8') as f:
        conteudo = f.read()
    for marcador, novo_bloco in blocos.items():
        padrao = f'/\\* {marcador}_START \\*/.*?/\\* {marcador}_END \\*/'
        if not re.search(padrao, conteudo, re.DOTALL):
            raise ValueError(f'[ERRO] Marcadores {marcador} nao encontrados no index.html.')
        conteudo = re.sub(padrao, novo_bloco, conteudo, flags=re.DOTALL)
    with open(index_path, 'w', encoding='utf-8') as f:
        f.write(conteudo)


# ═══════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════

def main():
    print()
    print('=' * 50)
    print('  ATUALIZADOR — Prospecção IEL')
    print('=' * 50)
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    try:
        STATUS_MAP.update(ler_depara(DEPARA_PATH))
        base = calcular_base()
        disc = calcular_discagem(encontrar_unico_xlsx(PASTA_DISCAGEM))
        bloco = {'IEL': gerar_bloco(base, disc)}
        atualizar_html(INDEX_HTML, bloco)
        print()
        print('=' * 50)
        print('  CONCLUIDO! index.html atualizado.')
        print('  Rode publicar.bat para enviar ao GitHub.')
        print('=' * 50)
        print()
    except Exception as e:
        print(f'\n[ERRO] {e}')
        import traceback; traceback.print_exc()


if __name__ == '__main__':
    main()
